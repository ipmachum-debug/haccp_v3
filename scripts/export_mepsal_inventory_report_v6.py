#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
멥쌀 재고원장 v6 생성 스크립트

v5와의 차이점:
  - 일별 원장 시트에 "제품 생산량(kg)" 컬럼 추가
  - 내용 컬럼을 "{제품}:{멥쌀_kg}kg → {제품}:{생산량}kg" 형식으로 확장

데이터 소스:
  1) 분배 상세 시트(v5)의 (일자, 제품명, 생산량) 매핑
  2) DB h_batches 테이블의 actual_quantity (tenant_id=2, 멥쌀 사용 배치)
  3) BOM% 역산: 멥쌀_사용량 / BOM_PCT × 100 = 추정_생산량 (위 두 가지 미존재 시 fallback)

출력:
  /root/haccp_v3/멥쌀_재고원장_v6.xlsx
"""
import os
import re
import sys
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    import pymysql
except ImportError:
    pymysql = None

SRC = "/root/haccp_v3/멥쌀_재고원장_v5.xlsx"
DST = "/root/haccp_v3/멥쌀_재고원장_v6.xlsx"

# BOM 멥쌀 함량(%) — 분배 상세 시트에서 추출되는 값과 동일 (fallback용 기본값)
BOM_PCT = {
    "쑥판인절미": 14.3,
    "판인절미": 15.1,
    "순인절미": 15.1,
    "콩고물쑥떡": 9.4,
    "콩고물쑥떡(동부)": 9.7,
    "마카다미아왕찹쌀떡(혼합)-흰": 9.7,
    "마카다미아왕찹쌀떡(혼합)-쑥": 9.4,
    # 일별 원장 주요 제품
    "쑥개떡": 15.1,           # 추정 (멥쌀가루 베이스)
    "모시개떡": 15.1,
    "꿀설기": 100.0,          # 설기는 멥쌀가루 100% 베이스 (멥쌀가루 자체)
    "습식 멥쌀가루": 100.0,    # 멥쌀가루는 멥쌀 100%
}

DB_CFG = {
    "host": "127.0.0.1", "port": 3306,
    "user": "root", "password": "G0ld3n!T1004#Sec",
    "database": "haccp_tenant_db", "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor if pymysql else None,
}


def load_distribution_map(wb):
    """분배 상세 시트에서 (일자, 제품명) → 생산량(kg) 매핑 추출."""
    m = {}
    if "분배 상세 (v5 신규)" not in wb.sheetnames:
        return m
    ws = wb["분배 상세 (v5 신규)"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_, name, prod_qty, bom_pct, mepsal_kg, scale, note = row[:7]
        if not date_ or name in (None, "합계"):
            continue
        key = (str(date_), str(name).strip())
        # 같은 날 같은 제품이 여러 배치면 합산
        m[key] = m.get(key, 0) + (float(prod_qty) if prod_qty else 0)
        # BOM% 도 보강
        if bom_pct and isinstance(bom_pct, (int, float)) and bom_pct > 0:
            BOM_PCT[str(name).strip()] = float(bom_pct)
    return m


def load_db_batch_map():
    """DB h_batches에서 (일자, 제품명) → actual_quantity 매핑.
    멥쌀(material_id=615) 을 사용한 배치만 대상.
    """
    if pymysql is None:
        print("  [WARN] pymysql 미설치 → DB 조회 생략")
        return {}
    try:
        conn = pymysql.connect(**DB_CFG)
    except Exception as e:
        print(f"  [WARN] DB 연결 실패 → DB 조회 생략: {e}")
        return {}
    m = {}
    sql = """
        SELECT DATE_FORMAT(b.planned_date, '%%Y-%%m-%%d') AS d,
               COALESCE(p2.product_name, p1.product_name, '?') AS name,
               COALESCE(b.actual_quantity, b.planned_quantity, 0) AS qty
        FROM h_batches b
        LEFT JOIN h_products_v2 p2 ON p2.id = b.product_id AND p2.tenant_id = b.tenant_id
        LEFT JOIN h_products    p1 ON p1.id = b.product_id AND p1.tenant_id = b.tenant_id
        WHERE b.tenant_id = 2
          AND b.id IN (
              SELECT DISTINCT reference_id
              FROM h_inventory_transactions
              WHERE tenant_id = 2 AND material_id = 615
                AND transaction_type IN ('usage','outbound')
                AND reference_type = 'batch'
          )
    """
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            for r in cur.fetchall():
                key = (r["d"], r["name"].strip())
                m[key] = m.get(key, 0) + float(r["qty"] or 0)
    except Exception as e:
        print(f"  [WARN] DB 조회 오류: {e}")
    finally:
        conn.close()
    return m


def calc_production_qty(date_str, product_name, mepsal_kg, dist_map, db_map):
    """우선순위: 분배상세 → DB → BOM% 역산."""
    key = (date_str, product_name)
    src = None
    qty = 0.0
    if key in dist_map and dist_map[key] > 0:
        qty = dist_map[key]
        src = "분배"
    elif key in db_map and db_map[key] > 0:
        qty = db_map[key]
        src = "DB"
    else:
        pct = BOM_PCT.get(product_name)
        if pct and pct > 0 and mepsal_kg:
            qty = round(mepsal_kg * 100.0 / pct, 1)
            src = f"BOM역산({pct}%)"
        else:
            src = "미상"
    return qty, src


def parse_content(content):
    """'쑥개떡:123.0kg, 모시개떡:82.0kg' → [('쑥개떡', 123.0), ('모시개떡', 82.0)]"""
    if not content:
        return []
    items = []
    for token in re.split(r"[,;]\s*", str(content)):
        token = token.strip()
        if not token:
            continue
        m = re.match(r"^(.+?)\s*:\s*([0-9.]+)\s*kg?\s*$", token, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            try:
                kg = float(m.group(2))
            except ValueError:
                continue
            items.append((name, kg))
    return items


def main():
    if not os.path.exists(SRC):
        print(f"ERROR: 원본 파일 없음: {SRC}", file=sys.stderr)
        sys.exit(1)

    print(f"[1] v5 로드: {SRC}")
    wb = load_workbook(SRC)

    print("[2] 분배 상세 시트 → 생산량 매핑 로드")
    dist_map = load_distribution_map(wb)
    print(f"    분배 매핑 {len(dist_map)}건")

    print("[3] DB h_batches → 실제 생산량 매핑 로드")
    db_map = load_db_batch_map()
    print(f"    DB 매핑 {len(db_map)}건")

    print("[4] 일별 원장 시트 변환")
    ws = wb["일별 원장"]

    # --- 헤더에 '제품 생산량(kg)' 컬럼 추가 (G열 '비고' 뒤에 H열로) ---
    # 기존: A=일자, B=구분, C=내용, D=입고, E=사용, F=잔량, G=비고
    # 신규: A=일자, B=구분, C=내용(확장), D=입고, E=사용, F=잔량, G=제품 생산량(kg), H=비고
    # 즉 G열에 새 컬럼 삽입
    ws.insert_cols(7)  # G열 자리에 새 빈 컬럼
    ws.cell(row=1, column=7, value="제품 생산량(kg)")

    # 헤더 스타일 일치
    hdr_a = ws.cell(row=1, column=1)
    new_hdr = ws.cell(row=1, column=7)
    if hdr_a.font:
        new_hdr.font = copy(hdr_a.font)
    if hdr_a.fill and hdr_a.fill.fgColor:
        new_hdr.fill = copy(hdr_a.fill)
    if hdr_a.alignment:
        new_hdr.alignment = copy(hdr_a.alignment)
    if hdr_a.border:
        new_hdr.border = copy(hdr_a.border)

    # 각 데이터 행에 대해 내용 확장 + 생산량 컬럼 채우기
    updated = 0
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        kind = ws.cell(row=r, column=2).value
        content = ws.cell(row=r, column=3).value
        usage = ws.cell(row=r, column=5).value

        if kind != "사용" or not content:
            continue

        date_str = str(date_cell) if date_cell else ""
        items = parse_content(content)
        if not items:
            continue

        new_parts = []
        total_prod = 0.0
        src_list = []
        for name, mepsal_kg in items:
            prod_qty, src = calc_production_qty(date_str, name, mepsal_kg, dist_map, db_map)
            total_prod += prod_qty
            src_list.append(src)
            if prod_qty > 0:
                new_parts.append(f"{name}:{mepsal_kg}kg→생산 {prod_qty:g}kg")
            else:
                new_parts.append(f"{name}:{mepsal_kg}kg→생산 ?")

        ws.cell(row=r, column=3, value=", ".join(new_parts))
        ws.cell(row=r, column=7, value=round(total_prod, 2) if total_prod else None)

        # 비고에 소스 정보 추가 (예: "정상 [분배,BOM역산]")
        note_cell = ws.cell(row=r, column=8)  # G가 새 컬럼이라 기존 비고는 H로 이동됨
        old_note = note_cell.value or ""
        # 같은 소스가 반복되면 한번만 표시
        uniq_src = ",".join(sorted(set(src_list)))
        if uniq_src and uniq_src not in old_note:
            note_cell.value = f"{old_note} [생산량:{uniq_src}]".strip()

        updated += 1

    print(f"    {updated} 행 업데이트 완료")

    # --- 열 폭 조정 ---
    widths = {
        "A": 12, "B": 8, "C": 60, "D": 10, "E": 10, "F": 11, "G": 14, "H": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # --- 요약 시트에 v6 메타 추가 ---
    if "요약" in wb.sheetnames:
        ws_sum = wb["요약"]
        last = ws_sum.max_row + 1
        ws_sum.cell(row=last, column=1, value="v6 보강")
        ws_sum.cell(row=last, column=2, value="제품 생산량(kg) 컬럼 추가")
        ws_sum.cell(row=last + 1, column=1, value="생성일")
        ws_sum.cell(row=last + 1, column=2, value="2026-06-30")

    print(f"[5] v6 저장: {DST}")
    wb.save(DST)
    print("    완료")
    print()
    print(f"  → 출력 파일: {DST}")


if __name__ == "__main__":
    main()
