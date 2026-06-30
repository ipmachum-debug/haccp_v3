#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
멥쌀 재고원장 v8 — 두 데이터소스 합집합으로 실제 배치 생산량 매칭

v7 한계: h_inventory_transactions(ref=batch) 만 사용 → 트랜잭션 ref가 NULL이거나
         batch_inputs로만 연결된 배치는 누락됨

v8 개선:
  매칭 소스 1: h_inventory_transactions (reference_type='batch', material_id=615)
  매칭 소스 2: h_batch_inputs (material_id=615)
  두 집합의 UNION을 (planned_date, product_name) 키로 합쳐 (멥쌀사용량, 제품생산량) 매핑

v5 분배상세 시트는 fallback (DB 미존재 행용)
"""
import os, re, sys
from copy import copy
import pymysql
from openpyxl import load_workbook

SRC = "/root/haccp_v3/멥쌀_재고원장_v5.xlsx"
DST = "/root/haccp_v3/멥쌀_재고원장_v8.xlsx"

DB_CFG = dict(host='127.0.0.1', port=3306, user='root', password='G0ld3n!T1004#Sec',
              database='haccp_tenant_db', charset='utf8mb4',
              cursorclass=pymysql.cursors.DictCursor)


def load_db_map():
    """두 소스(h_inventory_transactions + h_batch_inputs) UNION 으로 (일자,제품)→정보 매핑."""
    conn = pymysql.connect(**DB_CFG)
    cur = conn.cursor()

    # 멥쌀(615) 사용 배치를 찾는 두 가지 방법의 UNION
    # 같은 배치라도 reference 가 살아 있으면 source A, batch_inputs 만 있으면 source B
    sql = """
        SELECT
            CAST(b.planned_date AS CHAR) AS d,
            COALESCE(p2.product_name, p1.product_name, '?') AS pname,
            b.id AS batch_id,
            b.batch_code,
            b.actual_quantity AS product_qty,
            -- 멥쌀 사용량: batch_inputs 우선, 없으면 트랜잭션 합계
            COALESCE(bi.bi_qty, tx.tx_qty, 0) AS mepsal_used,
            b.status
        FROM h_batches b
        LEFT JOIN h_products_v2 p2 ON p2.id=b.product_id AND p2.tenant_id=b.tenant_id
        LEFT JOIN h_products    p1 ON p1.id=b.product_id AND p1.tenant_id=b.tenant_id
        LEFT JOIN (
            SELECT batch_id, SUM(actual_quantity) AS bi_qty
            FROM h_batch_inputs
            WHERE material_id=615
            GROUP BY batch_id
        ) bi ON bi.batch_id=b.id
        LEFT JOIN (
            SELECT reference_id AS bid, SUM(quantity) AS tx_qty
            FROM h_inventory_transactions
            WHERE tenant_id=2 AND material_id=615
              AND reference_type='batch'
              AND transaction_type IN ('usage','outbound')
            GROUP BY reference_id
        ) tx ON tx.bid=b.id
        WHERE b.tenant_id=2
          AND (bi.bi_qty IS NOT NULL OR tx.tx_qty IS NOT NULL)
    """
    cur.execute(sql)

    # (일자, 제품) → 멥쌀 사용량/생산량 집계 (여러 배치면 합산)
    m = {}
    for r in cur.fetchall():
        key = (str(r['d'])[:10], r['pname'].strip())
        if key not in m:
            m[key] = {'mepsal': 0.0, 'product': 0.0, 'codes': [], 'cnt': 0}
        m[key]['mepsal']  += float(r['mepsal_used'] or 0)
        m[key]['product'] += float(r['product_qty'] or 0)
        m[key]['codes'].append(r['batch_code'])
        m[key]['cnt'] += 1

    conn.close()
    return m


def parse_content(content):
    """'쑥개떡:123.0kg, 모시개떡:82.0kg' → [('쑥개떡', 123.0), ...]"""
    if not content:
        return []
    items = []
    for token in re.split(r"[,;]\s*", str(content)):
        token = token.strip()
        if not token:
            continue
        m = re.match(r"^(.+?)\s*:\s*([0-9.]+)\s*kg?\s*$", token, re.IGNORECASE)
        if m:
            try:
                items.append((m.group(1).strip(), float(m.group(2))))
            except ValueError:
                pass
    return items


def main():
    if not os.path.exists(SRC):
        sys.exit(f"ERROR: {SRC} 없음")

    print(f"[1] v5 로드: {SRC}")
    wb = load_workbook(SRC)

    # 분배 상세 시트 fallback 매핑
    dist_map = {}
    if "분배 상세 (v5 신규)" in wb.sheetnames:
        ws_d = wb["분배 상세 (v5 신규)"]
        for row in ws_d.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[1] in (None, "합계"):
                continue
            d, name, prod_q = str(row[0])[:10], str(row[1]).strip(), row[2]
            if prod_q:
                dist_map[(d, name)] = dist_map.get((d, name), 0) + float(prod_q)
    print(f"    분배상세 보조 매핑 {len(dist_map)}건")

    print("[2] DB 매핑 로드 (batch_inputs ∪ inventory_transactions)")
    db_map = load_db_map()
    print(f"    DB (일자,제품) 매핑 {len(db_map)}건")
    print("    샘플:")
    for k in sorted(db_map.keys())[:5]:
        v = db_map[k]
        print(f"      {k[0]} {k[1]:20s}  멥쌀={v['mepsal']:>7.2f}  생산={v['product']:>7.2f}  ({v['cnt']}배치)")

    print("[3] 일별 원장 시트 변환")
    ws = wb["일별 원장"]
    ws.insert_cols(7)
    ws.cell(row=1, column=7, value="제품 생산량(kg)")

    # 헤더 스타일
    hdr_a = ws.cell(row=1, column=1)
    new_hdr = ws.cell(row=1, column=7)
    if hdr_a.font:      new_hdr.font = copy(hdr_a.font)
    if hdr_a.fill:      new_hdr.fill = copy(hdr_a.fill)
    if hdr_a.alignment: new_hdr.alignment = copy(hdr_a.alignment)
    if hdr_a.border:    new_hdr.border = copy(hdr_a.border)

    matched_rows = 0
    matched_items = 0
    unmatched_items = []

    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        kind = ws.cell(row=r, column=2).value
        content = ws.cell(row=r, column=3).value
        if kind != "사용" or not content:
            continue

        date_str = str(date_cell)[:10] if date_cell else ""
        items = parse_content(content)
        if not items:
            continue

        new_parts = []
        total_prod = 0.0
        batch_codes = []
        sources = []
        row_has_match = False

        for name, mepsal_kg in items:
            info = db_map.get((date_str, name))
            if info and info['product'] > 0:
                prod = info['product']
                total_prod += prod
                batch_codes.extend(info['codes'])
                new_parts.append(f"{name}:{mepsal_kg}kg→생산 {prod:g}kg")
                sources.append("DB")
                row_has_match = True
                matched_items += 1
            else:
                # 분배상세 fallback
                fb = dist_map.get((date_str, name))
                if fb and fb > 0:
                    total_prod += fb
                    new_parts.append(f"{name}:{mepsal_kg}kg→생산 {fb:g}kg")
                    sources.append("v5분배")
                    row_has_match = True
                    matched_items += 1
                else:
                    new_parts.append(f"{name}:{mepsal_kg}kg→생산 ?")
                    unmatched_items.append((date_str, name, mepsal_kg))

        ws.cell(row=r, column=3, value=", ".join(new_parts))
        if total_prod > 0:
            ws.cell(row=r, column=7, value=round(total_prod, 2))
            matched_rows += 1

        # 비고 (열 H로 밀려있음)
        note_cell = ws.cell(row=r, column=8)
        old_note = note_cell.value or ""
        if row_has_match:
            tag_parts = []
            if batch_codes:
                tag_parts.append(f"배치:{','.join(c for c in batch_codes if c)}")
            uniq_src = ",".join(sorted(set(sources)))
            if uniq_src:
                tag_parts.append(f"출처:{uniq_src}")
            if tag_parts:
                tag = " [" + " | ".join(tag_parts) + "]"
                if tag.strip() not in old_note:
                    note_cell.value = f"{old_note}{tag}".strip()

    print(f"    매칭 성공 행: {matched_rows}, 매칭 성공 항목: {matched_items}")
    if unmatched_items:
        print(f"    매칭 실패 항목: {len(unmatched_items)}건")
        for u in unmatched_items[:15]:
            print(f"      - {u[0]} {u[1]} (멥쌀 {u[2]}kg)")
        if len(unmatched_items) > 15:
            print(f"      ... 외 {len(unmatched_items)-15}건")

    # 열 폭
    for col, w in {"A":12,"B":8,"C":80,"D":10,"E":10,"F":11,"G":15,"H":50}.items():
        ws.column_dimensions[col].width = w

    # 요약 시트 메타
    if "요약" in wb.sheetnames:
        ws_sum = wb["요약"]
        last = ws_sum.max_row + 1
        ws_sum.cell(row=last,   column=1, value="v8 보강")
        ws_sum.cell(row=last,   column=2, value="batch_inputs ∪ inventory_transactions 합집합")
        ws_sum.cell(row=last+1, column=1, value="생성일")
        ws_sum.cell(row=last+1, column=2, value="2026-06-30")
        ws_sum.cell(row=last+2, column=1, value="매칭 성공 행")
        ws_sum.cell(row=last+2, column=2, value=matched_rows)
        ws_sum.cell(row=last+3, column=1, value="매칭 실패 항목")
        ws_sum.cell(row=last+3, column=2, value=len(unmatched_items))

    print(f"[4] v8 저장: {DST}")
    wb.save(DST)
    print("    완료")


if __name__ == "__main__":
    main()
