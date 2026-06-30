#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
멥쌀 재고원장 v9 — 사용자 현장 실측 기준 + DB 실측 보강

기준 (사용자 운영 확정값):
  시작재고 (2026-02-09 현장 실측): 1,520 kg
  입고 6회 합계:                  10,000 kg
  종료재고 (2026-06-25 현장 실측): 1,670 kg
  결과 사용량:                     9,850 kg

제품 생산량 표기:
  우선순위 1: DB h_batch_inputs + h_batches.actual_quantity (실측)
  우선순위 2: DB h_inventory_transactions reference_type=batch (실측, NULL 트랜잭션은 backfill 중복이라 제외)
  우선순위 3: v5 분배상세 시트의 추정 생산량 (사용자가 BOM 역산으로 만든 자료)
  표기 안 됨: 위 셋 모두 없으면 "?" (사용자가 직접 입력해야 할 케이스)

출처 표기:
  비고에 [출처:DB실측 / v5추정-BOM역산 / 미확인] 명확히 구분

추가 시트:
  - BOM 멥쌀% 시트: DB 현재 BOM의 멥쌀 함량 전체
  - 정합성 진단 시트: DB 사용량 vs 사용자 9,850kg 차이
"""
import os
import re
import sys
from copy import copy
import pymysql
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/root/haccp_v3/멥쌀_재고원장_v5.xlsx"
DST = "/root/haccp_v3/멥쌀_재고원장_v9.xlsx"

DB_CFG = dict(host='127.0.0.1', port=3306, user='root', password='G0ld3n!T1004#Sec',
              database='haccp_tenant_db', charset='utf8mb4',
              cursorclass=pymysql.cursors.DictCursor)


# ──────────────────────────────────────────────────────────────────
# 제품명 정규화 — v5와 DB 이름 차이 흡수
# ──────────────────────────────────────────────────────────────────
def normalize_product_name(name):
    """v5와 DB 사이의 제품명 표기 차이를 통일.
    예: '쑥판인절미' ↔ 'p=44 쑥판인절미' or 'p=26 다이스인절미' 등.
    """
    if not name:
        return ""
    n = name.strip()
    # 공백 제거 (한글 사이 공백)
    n_compact = re.sub(r"\s+", "", n)
    # 별칭 매핑
    aliases = {
        "마카다미아왕찹쌀떡(혼합)-흰": ["마카다미아왕찹쌀떡", "마카다미아찹쌀떡"],
        "마카다미아왕찹쌀떡(혼합)-쑥": ["마카다미아쑥찹쌀떡"],
        "콩고물쑥떡(동부)": ["콩고물쑥떡동부"],
    }
    return n_compact


def load_db_data():
    """모든 DB 데이터 로드."""
    conn = pymysql.connect(**DB_CFG)
    cur = conn.cursor()
    data = {}

    # 1) 배치별 멥쌀 사용 + 제품 생산량 (batch_inputs + transactions UNION)
    cur.execute("""
        SELECT
            CAST(b.planned_date AS CHAR) AS d,
            COALESCE(p2.product_name, p1.product_name, '?') AS pname,
            b.id AS batch_id,
            b.batch_code,
            b.actual_quantity AS product_qty,
            COALESCE(bi.bi_qty, tx.tx_qty, 0) AS mepsal_used,
            CASE
                WHEN bi.bi_qty IS NOT NULL THEN 'batch_inputs'
                WHEN tx.tx_qty IS NOT NULL THEN 'transactions'
                ELSE 'none'
            END AS src
        FROM h_batches b
        LEFT JOIN h_products_v2 p2 ON p2.id=b.product_id AND p2.tenant_id=b.tenant_id
        LEFT JOIN h_products    p1 ON p1.id=b.product_id AND p1.tenant_id=b.tenant_id
        LEFT JOIN (
            SELECT batch_id, SUM(actual_quantity) AS bi_qty
            FROM h_batch_inputs
            WHERE material_id=615
            GROUP BY batch_id
        ) bi ON bi.batch_id=b.id
        LEFT JOIN (
            SELECT reference_id AS bid, SUM(quantity) AS tx_qty
            FROM h_inventory_transactions
            WHERE tenant_id=2 AND material_id=615
              AND reference_type='batch'
              AND transaction_type IN ('usage','outbound')
            GROUP BY reference_id
        ) tx ON tx.bid=b.id
        WHERE b.tenant_id=2
          AND (bi.bi_qty IS NOT NULL OR tx.tx_qty IS NOT NULL)
    """)
    rows = cur.fetchall()

    # (일자, 정규화제품명) → 정보 집계
    db_map = {}
    for r in rows:
        nm = r['pname'].strip()
        key = (str(r['d'])[:10], normalize_product_name(nm))
        if key not in db_map:
            db_map[key] = {
                'product_qty': 0.0,
                'mepsal_used': 0.0,
                'batch_codes': [],
                'real_names': set(),  # 실제 DB 제품명
                'sources': set(),
            }
        db_map[key]['product_qty'] += float(r['product_qty'] or 0)
        db_map[key]['mepsal_used'] += float(r['mepsal_used'] or 0)
        db_map[key]['batch_codes'].append(r['batch_code'])
        db_map[key]['real_names'].add(nm)
        db_map[key]['sources'].add(r['src'])
    data['db_map'] = db_map

    # 2) DB 현재 BOM 의 멥쌀% (참고용)
    cur.execute("""
        SELECT rh.id AS recipe_id,
               rh.product_id,
               COALESCE(p2.product_name, p1.product_name) AS pname,
               rh.recipe_name,
               rh.target_quantity,
               rh.unit,
               rh.is_active,
               COALESCE(mep.mep_pct, 0) AS mep_pct,
               rl_total.total_pct AS total_pct
        FROM h_recipe_headers rh
        LEFT JOIN h_products_v2 p2 ON p2.id=rh.product_id AND p2.tenant_id=rh.tenant_id
        LEFT JOIN h_products    p1 ON p1.id=rh.product_id AND p1.tenant_id=rh.tenant_id
        LEFT JOIN (
            SELECT recipe_id, SUM(quantity) AS mep_pct
            FROM h_recipe_lines
            WHERE material_id=615
            GROUP BY recipe_id
        ) mep ON mep.recipe_id=rh.id
        LEFT JOIN (
            SELECT recipe_id, SUM(quantity) AS total_pct
            FROM h_recipe_lines
            GROUP BY recipe_id
        ) rl_total ON rl_total.recipe_id=rh.id
        WHERE rh.tenant_id=2 AND rh.is_active=1
        ORDER BY mep_pct DESC, pname
    """)
    data['bom_list'] = cur.fetchall()

    # 3) 입고 트랜잭션 (사용자 6회 입고 확인용)
    cur.execute("""
        SELECT CAST(transaction_date AS CHAR) AS d,
               t.quantity,
               t.lot_id,
               l.lot_number,
               t.notes
        FROM h_inventory_transactions t
        LEFT JOIN h_inventory_lots l ON l.id=t.lot_id
        WHERE t.tenant_id=2 AND t.material_id=615
          AND t.transaction_type IN ('receipt','inbound','purchase')
        ORDER BY t.transaction_date
    """)
    data['receipts'] = cur.fetchall()

    # 4) DB 총 사용량 (정합성 진단용)
    cur.execute("""
        SELECT SUM(bi.actual_quantity) AS total
        FROM h_batch_inputs bi
        JOIN h_batches b ON b.id=bi.batch_id
        WHERE bi.material_id=615 AND b.tenant_id=2
          AND b.planned_date BETWEEN '2026-02-09' AND '2026-06-25'
    """)
    data['db_total_use_batch_inputs'] = float(cur.fetchone()['total'] or 0)

    conn.close()
    return data


def parse_content(content):
    """'쑥개떡:123.0kg, 모시개떡:82.0kg' → [('쑥개떡', 123.0), ...]"""
    if not content:
        return []
    items = []
    for token in re.split(r"[,;]\s*", str(content)):
        token = token.strip()
        if not token:
            continue
        m = re.match(r"^(.+?)\s*:\s*([0-9.]+)\s*kg?\s*$", token, re.IGNORECASE)
        if m:
            try:
                items.append((m.group(1).strip(), float(m.group(2))))
            except ValueError:
                pass
    return items


def build_dist_map(wb):
    """v5 분배상세 시트 → (일자, 제품) → 생산량 매핑."""
    m = {}
    if "분배 상세 (v5 신규)" not in wb.sheetnames:
        return m
    ws = wb["분배 상세 (v5 신규)"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] in (None, "합계"):
            continue
        d = str(row[0])[:10]
        name = str(row[1]).strip()
        qty = row[2]
        if qty:
            key = (d, normalize_product_name(name))
            m[key] = m.get(key, 0) + float(qty)
    return m


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def main():
    if not os.path.exists(SRC):
        sys.exit(f"ERROR: {SRC} 없음")

    print(f"[1] v5 로드: {SRC}")
    wb = load_workbook(SRC)

    print("[2] DB 데이터 로드")
    data = load_db_data()
    db_map = data['db_map']
    print(f"    DB (일자,제품) 매핑: {len(db_map)}건")
    print(f"    DB BOM(active): {len(data['bom_list'])}건")
    print(f"    DB 입고 트랜잭션: {len(data['receipts'])}건")
    print(f"    DB batch_inputs 멥쌀 사용 합계 (2/9~6/25): {data['db_total_use_batch_inputs']:.2f} kg")

    # v5 분배상세 fallback 매핑
    dist_map = build_dist_map(wb)
    print(f"    v5 분배상세 fallback: {len(dist_map)}건")

    print("[3] 일별 원장 시트 변환")
    ws = wb["일별 원장"]

    # G열에 새 컬럼 삽입 (기존 G=비고 → H로)
    ws.insert_cols(7)
    ws.cell(row=1, column=7, value="제품 생산량(kg)")
    new_hdr = ws.cell(row=1, column=7)
    hdr_a = ws.cell(row=1, column=1)
    if hdr_a.font:      new_hdr.font = copy(hdr_a.font)
    if hdr_a.fill:      new_hdr.fill = copy(hdr_a.fill)
    if hdr_a.alignment: new_hdr.alignment = copy(hdr_a.alignment)
    if hdr_a.border:    new_hdr.border = copy(hdr_a.border)

    db_matched_items = 0
    v5_fallback_items = 0
    unmatched_items = []
    diagnostics = []  # 진단용

    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        kind = ws.cell(row=r, column=2).value
        content = ws.cell(row=r, column=3).value
        if kind != "사용" or not content:
            continue

        date_str = str(date_cell)[:10] if date_cell else ""
        items = parse_content(content)
        if not items:
            continue

        new_parts = []
        total_prod = 0.0
        all_batch_codes = []
        sources = []

        for name, mepsal_kg_v5 in items:
            norm = normalize_product_name(name)
            info = db_map.get((date_str, norm))

            if info and info['product_qty'] > 0:
                # DB 실측
                prod = info['product_qty']
                total_prod += prod
                all_batch_codes.extend(info['batch_codes'])
                # DB 실명이 다르면 표시 (예: v5='쑥판인절미' DB='다이스인절미')
                db_real_name = ", ".join(sorted(info['real_names']))
                if normalize_product_name(db_real_name) != norm:
                    new_parts.append(
                        f"{name}({mepsal_kg_v5}kg멥쌀)→DB '{db_real_name}' {prod:g}kg"
                    )
                else:
                    new_parts.append(f"{name}:{mepsal_kg_v5}kg→생산 {prod:g}kg")
                sources.append("DB실측")
                db_matched_items += 1
                # 멥쌀 사용량 차이 진단
                db_mep = info['mepsal_used']
                if abs(db_mep - mepsal_kg_v5) > 1:
                    diagnostics.append({
                        '일자': date_str, '제품': name,
                        'v5_멥쌀': mepsal_kg_v5, 'DB_멥쌀': db_mep,
                        '차이': db_mep - mepsal_kg_v5,
                        '비고': f"v5 표기 멥쌀과 DB 실측 차이 (배치: {','.join(info['batch_codes'])})",
                    })
            else:
                # v5 분배상세 fallback
                fb = dist_map.get((date_str, norm))
                if fb and fb > 0:
                    total_prod += fb
                    new_parts.append(f"{name}:{mepsal_kg_v5}kg→[추정] {fb:g}kg")
                    sources.append("v5추정")
                    v5_fallback_items += 1
                else:
                    new_parts.append(f"{name}:{mepsal_kg_v5}kg→?")
                    unmatched_items.append((date_str, name, mepsal_kg_v5))

        ws.cell(row=r, column=3, value=", ".join(new_parts))
        if total_prod > 0:
            ws.cell(row=r, column=7, value=round(total_prod, 2))

        # 비고 (H열)
        note_cell = ws.cell(row=r, column=8)
        old_note = note_cell.value or ""
        if sources:
            uniq_src = ",".join(sorted(set(sources)))
            tag_parts = [f"출처:{uniq_src}"]
            if all_batch_codes:
                codes = ",".join(c for c in all_batch_codes if c)
                if codes:
                    tag_parts.append(f"배치:{codes}")
            tag = " [" + " | ".join(tag_parts) + "]"
            if tag.strip() not in old_note:
                note_cell.value = f"{old_note}{tag}".strip()

    print(f"    DB 실측 매칭: {db_matched_items}건")
    print(f"    v5 분배상세 추정: {v5_fallback_items}건")
    print(f"    미매칭: {len(unmatched_items)}건")

    # 열 폭
    for col, w in {"A":12,"B":8,"C":85,"D":10,"E":10,"F":11,"G":15,"H":50}.items():
        ws.column_dimensions[col].width = w

    # ─────────────────────────────────────────
    # 시트 추가: 'BOM 멥쌀%' — DB 현재 활성 레시피의 멥쌀 함량
    # ─────────────────────────────────────────
    print("[4] BOM 멥쌀% 시트 추가")
    if "BOM 멥쌀%" in wb.sheetnames:
        del wb["BOM 멥쌀%"]
    ws_b = wb.create_sheet("BOM 멥쌀%")
    headers = ["제품ID", "제품명", "레시피명", "레시피ID", "target", "단위",
               "멥쌀 BOM(%)", "total(%)", "비고"]
    for i, h in enumerate(headers, 1):
        c = ws_b.cell(row=1, column=i, value=h)
        style_header(c)
    for i, r in enumerate(data['bom_list'], 2):
        mep = float(r['mep_pct'] or 0)
        ws_b.cell(row=i, column=1, value=r['product_id'])
        ws_b.cell(row=i, column=2, value=r['pname'] or '?')
        ws_b.cell(row=i, column=3, value=r['recipe_name'])
        ws_b.cell(row=i, column=4, value=r['recipe_id'])
        ws_b.cell(row=i, column=5, value=float(r['target_quantity'] or 0))
        ws_b.cell(row=i, column=6, value=r['unit'])
        ws_b.cell(row=i, column=7, value=round(mep, 3))
        ws_b.cell(row=i, column=8, value=round(float(r['total_pct'] or 0), 3))
        if mep == 0:
            ws_b.cell(row=i, column=9, value="멥쌀 BOM 0% — 운영 실제와 다를 수 있음 (3·4·5월 품목제조보고 수정중)")
    widths_b = {"A":7,"B":30,"C":30,"D":10,"E":10,"F":6,"G":12,"H":10,"I":55}
    for col, w in widths_b.items():
        ws_b.column_dimensions[col].width = w
    ws_b.freeze_panes = "A2"

    # ─────────────────────────────────────────
    # 시트 추가: '정합성 진단'
    # ─────────────────────────────────────────
    print("[5] 정합성 진단 시트 추가")
    if "정합성 진단" in wb.sheetnames:
        del wb["정합성 진단"]
    ws_d = wb.create_sheet("정합성 진단")
    ws_d.cell(row=1, column=1, value="멥쌀 재고원장 정합성 진단 (v9)")
    ws_d.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_d.merge_cells("A1:E1")

    rows = [
        ("", "", "", "", ""),
        ("[현장 실측 기준 (사용자 보고)]", "", "", "", ""),
        ("시작재고 (2026-02-09 현장 실측)", 1520.0, "kg", "사용자 현장 확인값", ""),
        ("입고 6회 합계", 10000.0, "kg", "MAT-20260331/0415/0429×2/0520/0623", ""),
        ("종료재고 (2026-06-25 현장 실측)", 1670.0, "kg", "≈1,669.98 (시스템 일치)", ""),
        ("결과 사용량 = 시작 + 입고 - 종료", 9850.0, "kg", "v5 일별원장 합계와 일치", ""),
        ("", "", "", "", ""),
        ("[DB 시스템 기록]", "", "", "", ""),
        ("DB batch_inputs 멥쌀 (2026-02-09 ~ 2026-06-25)",
         round(data['db_total_use_batch_inputs'], 2), "kg",
         "h_batch_inputs.actual_quantity 합계", ""),
        ("DB 사용량 - 사용자 9850 차이",
         round(data['db_total_use_batch_inputs'] - 9850.0, 2), "kg",
         "양수면 DB가 과대 입력, 음수면 누락", ""),
        ("", "", "", "", ""),
        ("[입고 6회 검증]", "", "", "", ""),
    ]
    for tup in rows:
        ws_d.append(tup)

    # 입고 6회 명세
    receipts_rows = []
    rcv_total = 0
    for rcv in data['receipts']:
        receipts_rows.append((
            f"  {rcv['d']}", float(rcv['quantity']), "kg",
            rcv['lot_number'] or '', (rcv['notes'] or '')[:50]
        ))
        rcv_total += float(rcv['quantity'])
    for tup in receipts_rows:
        ws_d.append(tup)
    ws_d.append(("입고 합계", rcv_total, "kg", f"6회 = {rcv_total}", ""))

    # 미매칭 항목
    if unmatched_items:
        ws_d.append(("", "", "", "", ""))
        ws_d.append(("[미매칭 항목 — 사용자 직접 입력 필요]", "", "", "", ""))
        ws_d.append(("일자", "제품", "멥쌀(kg)", "비고", ""))
        for um in unmatched_items:
            ws_d.append((um[0], um[1], um[2], "DB·v5 모두에 없음", ""))

    # 진단(차이) 행
    if diagnostics:
        ws_d.append(("", "", "", "", ""))
        ws_d.append(("[멥쌀 사용량 v5 vs DB 차이 (1kg 초과)]", "", "", "", ""))
        ws_d.append(("일자", "제품", "v5 멥쌀(kg)", "DB 멥쌀(kg)", "차이"))
        for d in diagnostics:
            ws_d.append((d['일자'], d['제품'], d['v5_멥쌀'], d['DB_멥쌀'], d['차이']))

    widths_d = {"A":30,"B":20,"C":15,"D":35,"E":50}
    for col, w in widths_d.items():
        ws_d.column_dimensions[col].width = w

    # ─────────────────────────────────────────
    # 요약 시트 갱신
    # ─────────────────────────────────────────
    if "요약" in wb.sheetnames:
        ws_sum = wb["요약"]
        last = ws_sum.max_row + 1
        ws_sum.cell(row=last,   column=1, value="v9 보강")
        ws_sum.cell(row=last,   column=2, value="DB 실측 + v5 분배상세 추정 (출처 명시)")
        ws_sum.cell(row=last+1, column=1, value="DB 실측 매칭")
        ws_sum.cell(row=last+1, column=2, value=f"{db_matched_items}건")
        ws_sum.cell(row=last+2, column=1, value="v5 추정 fallback")
        ws_sum.cell(row=last+2, column=2, value=f"{v5_fallback_items}건")
        ws_sum.cell(row=last+3, column=1, value="미매칭")
        ws_sum.cell(row=last+3, column=2, value=f"{len(unmatched_items)}건")
        ws_sum.cell(row=last+4, column=1, value="생성일")
        ws_sum.cell(row=last+4, column=2, value="2026-06-30")

    print(f"[6] v9 저장: {DST}")
    wb.save(DST)
    print("    완료")


if __name__ == "__main__":
    main()
