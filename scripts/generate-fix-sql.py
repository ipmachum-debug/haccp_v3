#!/usr/bin/env python3
"""
Excel 기반 수불부 데이터 보정 SQL 생성기

엑셀의 📊 월별 원료수불부 시트에서 전월재고를 읽고,
material_ledger_daily의 running_stock을 보정하는 SQL을 생성합니다.

Usage: python3 scripts/generate-fix-sql.py > scripts/fix-data.sql
"""
import openpyxl
import json
from datetime import datetime

TENANT_ID = 2
EXCEL_PATH = 'HACCP_원료수불부_원가관리0320.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ── 1. 원료 마스터에서 이름 매핑 (엑셀 이름 → DB에서 매칭할 이름)
ws_master = wb['🏭 원료 마스터']
material_names = []
for row in ws_master.iter_rows(min_row=4, max_row=ws_master.max_row, values_only=True):
    if row[0] and row[1]:
        material_names.append(str(row[1]).strip())

# ── 2. 월별 수불부에서 전월재고 (2026-01 기준) 추출
ws_ledger = wb['📊 월별 원료수불부']
period = str(ws_ledger.cell(row=2, column=2).value)  # '2026-01'
print(f"-- Excel period: {period}")
print(f"-- Tenant: {TENANT_ID}")
print(f"-- Generated: {datetime.now().isoformat()}")
print()

# 전월재고 데이터 (이름 → 수량)
initial_stocks = {}
for row in ws_ledger.iter_rows(min_row=5, max_row=ws_ledger.max_row, values_only=True):
    if row[1]:
        name = str(row[1]).strip()
        prev_stock = float(row[2] or 0)  # 전월재고
        initial_stocks[name] = prev_stock

print(f"-- Materials with initial stock: {sum(1 for v in initial_stocks.values() if v > 0)}")
print()

# ── 3. 입고 데이터 추출 (📥 원재료 입고)
ws_recv = wb['📥 원재료 입고']
receiving_records = []
for row in ws_recv.iter_rows(min_row=6, max_row=ws_recv.max_row, values_only=True):
    if row[0] and isinstance(row[0], datetime):
        date_str = row[0].strftime('%Y-%m-%d')
        material = str(row[1]).strip() if row[1] else ''
        partner = str(row[2]).strip() if row[2] else ''
        recv_type = str(row[3]).strip() if row[3] else ''  # 자체구매/위탁공급
        qty = float(row[4]) if row[4] else 0
        unit_price = float(row[5]) if row[5] else 0
        amount = float(row[6]) if row[6] else 0
        expiry = row[8].strftime('%Y-%m-%d') if isinstance(row[8], datetime) else ''
        note = str(row[9]).strip() if row[9] else ''
        
        if material and qty > 0:
            receiving_records.append({
                'date': date_str,
                'material': material,
                'partner': partner,
                'type': recv_type,
                'qty': qty,
                'unit_price': unit_price,
                'amount': amount,
                'expiry': expiry,
                'note': note
            })

print(f"-- Receiving records from Excel: {len(receiving_records)}")
print()

# ── 4. SQL 생성 ──

print("-- ============================================")
print("-- STEP 0: 기존 seed 소스의 잘못된 데이터 정리")
print("-- ============================================")
print()

# material_ledger_daily에 receiving data가 없는 경우만 처리
# (기존 auto_purchase 소스 데이터는 유지)
print("-- 이전 시드 데이터의 running_stock을 리셋 (나중에 재계산)")
print(f"UPDATE material_ledger_daily SET running_stock = 0 WHERE tenant_id = {TENANT_ID};")
print()

print("-- ============================================")
print("-- STEP 1: 초기 재고 (전월재고) 설정")
print("-- 2025-12-31 기준 각 원재료의 시작 재고")
print("-- ============================================")
print()

# 초기 재고를 2025-12-31 날짜로 adjustment로 기록
for name, stock in initial_stocks.items():
    if stock > 0:
        safe_name = name.replace("'", "\\'")
        print(f"""-- {name}: 전월재고 {stock}kg
INSERT INTO material_ledger_daily (tenant_id, material_id, ledger_date, receiving_qty, usage_qty, adjustment_qty, running_stock, notes, source)
SELECT {TENANT_ID}, m.id, '2025-12-31', 0, 0, {stock:.3f}, {stock:.3f}, '엑셀 전월재고 이월', 'excel_initial'
FROM h_materials m WHERE m.tenant_id = {TENANT_ID} AND m.material_name = '{safe_name}'
ON DUPLICATE KEY UPDATE adjustment_qty = {stock:.3f}, running_stock = {stock:.3f}, notes = '엑셀 전월재고 이월', source = 'excel_initial';
""")

print()
print("-- ============================================")
print("-- STEP 2: 엑셀 입고 데이터를 material_ledger_daily에 반영")
print("-- ============================================")
print()

# 날짜+재료 별 입고합계
from collections import defaultdict
recv_by_date_material = defaultdict(float)
for r in receiving_records:
    key = (r['date'], r['material'])
    recv_by_date_material[key] += r['qty']

for (date, material), qty in sorted(recv_by_date_material.items()):
    safe_name = material.replace("'", "\\'")
    print(f"""-- {date} {material}: 입고 {qty}kg
INSERT INTO material_ledger_daily (tenant_id, material_id, ledger_date, receiving_qty, usage_qty, adjustment_qty, running_stock, notes, source)
SELECT {TENANT_ID}, m.id, '{date}', {qty:.3f}, 0, 0, 0, '엑셀 입고', 'excel_receiving'
FROM h_materials m WHERE m.tenant_id = {TENANT_ID} AND m.material_name = '{safe_name}'
ON DUPLICATE KEY UPDATE receiving_qty = receiving_qty + {qty:.3f}, notes = CONCAT(COALESCE(notes,''), ' +엑셀입고'), source = 'excel_receiving';
""")

print()
print("-- ============================================")
print("-- STEP 3: running_stock 재계산")
print("-- 원재료별 날짜순으로 누적 계산")
print("-- ============================================")
print()

# running_stock 재계산은 프로시저로 처리
print("""
-- running_stock 재계산 프로시저
DELIMITER //
DROP PROCEDURE IF EXISTS recalc_running_stock //
CREATE PROCEDURE recalc_running_stock(IN p_tenant_id INT)
BEGIN
    DECLARE done INT DEFAULT FALSE;
    DECLARE v_mat_id BIGINT;
    DECLARE v_running DECIMAL(12,3);
    DECLARE v_id BIGINT;
    DECLARE v_recv DECIMAL(12,3);
    DECLARE v_usage DECIMAL(12,3);
    DECLARE v_adj DECIMAL(12,3);
    
    -- 원재료 커서
    DECLARE mat_cursor CURSOR FOR
        SELECT DISTINCT material_id FROM material_ledger_daily WHERE tenant_id = p_tenant_id;
    DECLARE CONTINUE HANDLER FOR NOT FOUND SET done = TRUE;
    
    OPEN mat_cursor;
    mat_loop: LOOP
        FETCH mat_cursor INTO v_mat_id;
        IF done THEN LEAVE mat_loop; END IF;
        
        SET v_running = 0;
        
        -- 해당 원재료의 모든 일별 레코드를 날짜순으로 업데이트
        BEGIN
            DECLARE done2 INT DEFAULT FALSE;
            DECLARE day_cursor CURSOR FOR
                SELECT id, COALESCE(receiving_qty,0), COALESCE(usage_qty,0), COALESCE(adjustment_qty,0)
                FROM material_ledger_daily
                WHERE tenant_id = p_tenant_id AND material_id = v_mat_id
                ORDER BY ledger_date ASC, id ASC;
            DECLARE CONTINUE HANDLER FOR NOT FOUND SET done2 = TRUE;
            
            OPEN day_cursor;
            day_loop: LOOP
                FETCH day_cursor INTO v_id, v_recv, v_usage, v_adj;
                IF done2 THEN LEAVE day_loop; END IF;
                
                SET v_running = v_running + v_recv - v_usage + v_adj;
                UPDATE material_ledger_daily SET running_stock = v_running WHERE id = v_id;
            END LOOP day_loop;
            CLOSE day_cursor;
        END;
        
    END LOOP mat_loop;
    CLOSE mat_cursor;
    
    SELECT 'running_stock recalculation complete' as result;
END //
DELIMITER ;

CALL recalc_running_stock(""" + str(TENANT_ID) + """);
DROP PROCEDURE IF EXISTS recalc_running_stock;
""")

print()
print("-- ============================================")
print("-- STEP 4: h_inventory_transactions에 usage 레코드 생성")
print("-- (batch 기반)")
print("-- ============================================")
print()

print(f"""
-- 기존 usage 트랜잭션 삭제 (재생성)
DELETE FROM h_inventory_transactions WHERE tenant_id = {TENANT_ID} AND transaction_type = 'usage';

-- batch_inputs 기반으로 usage 트랜잭션 생성
-- lot_id가 NULL인 경우 해당 material의 첫 번째 available LOT 사용
INSERT INTO h_inventory_transactions
    (tenant_id, lot_id, transaction_type, quantity, unit, transaction_date,
     reference_type, reference_id, source_type, notes, created_by)
SELECT 
    bi.tenant_id,
    COALESCE(bi.lot_id, (
        SELECT il.id FROM h_inventory_lots il
        WHERE il.tenant_id = bi.tenant_id 
          AND il.material_id = bi.material_id 
          AND il.status = 'available'
        ORDER BY il.receipt_date ASC, il.id ASC
        LIMIT 1
    )),
    'usage',
    COALESCE(bi.actual_quantity, bi.planned_quantity),
    COALESCE(bi.unit, 'kg'),
    b.planned_date,
    'batch',
    bi.batch_id,
    'batch_completion',
    CONCAT('생산투입-배치#', bi.batch_id),
    1
FROM h_batch_inputs bi
JOIN h_batches b ON b.id = bi.batch_id AND b.tenant_id = bi.tenant_id
WHERE bi.tenant_id = {TENANT_ID}
  AND b.status = 'completed'
  AND COALESCE(bi.actual_quantity, bi.planned_quantity) > 0;
""")

print()
print("-- ============================================")
print("-- STEP 5: 엑셀의 입고 데이터를 h_inventory_lots에도 반영")
print("-- (기존 LOT에 없는 입고건만)")
print("-- ============================================")
print()

# 엑셀 입고 데이터에서 LOT 생성
for i, r in enumerate(receiving_records):
    safe_name = r['material'].replace("'", "\\'")
    safe_partner = r['partner'].replace("'", "\\'") if r['partner'] else ''
    lot_num = f"LOT-EXCEL-{r['date'].replace('-','')}-{i+1:04d}"
    
    print(f"""-- {r['date']} {r['material']} {r['qty']}kg from {r['partner']}
INSERT IGNORE INTO h_inventory_lots 
    (tenant_id, lot_number, material_id, quantity, current_quantity, available_quantity, 
     unit, unit_price, receipt_date, supplier_name, expiry_date, status)
SELECT {TENANT_ID}, '{lot_num}', m.id, {r['qty']:.3f}, {r['qty']:.3f}, {r['qty']:.3f},
       'kg', {r['unit_price']:.2f}, '{r['date']}', '{safe_partner}', 
       {f"'{r['expiry']}'" if r['expiry'] else 'NULL'}, 'available'
FROM h_materials m WHERE m.tenant_id = {TENANT_ID} AND m.material_name = '{safe_name}'
AND NOT EXISTS (
    SELECT 1 FROM h_inventory_lots il 
    WHERE il.tenant_id = {TENANT_ID} AND il.material_id = m.id 
      AND il.receipt_date = '{r['date']}' AND ABS(il.quantity - {r['qty']:.3f}) < 0.01
);
""")

print()
print("-- ============================================")
print("-- STEP 6: 입고 LOT에 대한 receipt 트랜잭션 생성")
print("-- ============================================")
print()

print(f"""
-- Excel LOT의 receipt 트랜잭션 (중복 방지)
INSERT INTO h_inventory_transactions
    (tenant_id, lot_id, transaction_type, quantity, unit, transaction_date,
     reference_type, source_type, notes, created_by)
SELECT 
    il.tenant_id,
    il.id,
    'receipt',
    il.quantity,
    il.unit,
    il.receipt_date,
    'excel_import',
    'excel_import',
    CONCAT('엑셀입고-', il.supplier_name),
    1
FROM h_inventory_lots il
WHERE il.tenant_id = {TENANT_ID}
  AND il.lot_number LIKE 'LOT-EXCEL-%'
  AND NOT EXISTS (
      SELECT 1 FROM h_inventory_transactions t
      WHERE t.tenant_id = il.tenant_id AND t.lot_id = il.id AND t.transaction_type = 'receipt'
  );
""")

print()
print("-- ============================================")
print("-- DONE")  
print("-- ============================================")
print(f"-- Total receiving records: {len(receiving_records)}")
print(f"-- Initial stock entries: {sum(1 for v in initial_stocks.values() if v > 0)}")
print()
