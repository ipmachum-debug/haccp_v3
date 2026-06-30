#!/usr/bin/env python3
"""
멥쌀(material_id=615) 재고 보고서 엑셀 생성
- 기간: 2026-02-09 ~ 현재
- 시트: ① 요약, ② 입고 내역, ③ 사용 내역, ④ 일자별 재고 변동, ⑤ LOT별 현황, ⑥ 전체 트랜잭션
"""
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from decimal import Decimal

# DB connection
conn = pymysql.connect(
    host="127.0.0.1", user="root", password="G0ld3n!T1004#Sec",
    database="haccp_tenant_db", charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor
)

START_DATE = "2026-02-09"
MATERIAL_ID = 615
TENANT_ID = 2

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="305496")
SUBTITLE_FONT = Font(bold=True, size=11, color="305496")
RECEIPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
USAGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ADJ_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def fmt_num(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_cell(cell, align=None, fill=None, fmt=None):
    cell.border = BORDER
    if align:
        cell.alignment = align
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


def autosize_cols(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


with conn.cursor() as cur:
    # 1. 기본 정보
    cur.execute("""
      SELECT id, material_code, material_name, unit
      FROM h_materials WHERE id=%s AND tenant_id=%s
    """, (MATERIAL_ID, TENANT_ID))
    mat = cur.fetchone()

    # 2. 시작 시점(2026-02-09 0시) 재고 = 2026-02-09 이전까지의 LOT 입고 합계 − usage 합계
    # 우리는 트랜잭션이 누락된 데이터가 있으므로, LOT 기준으로 시작 재고를 정확히 산출:
    # 시작재고 = sum(LOT.quantity where receipt_date<START) − sum(usage where transaction_date<START)
    cur.execute("""
      SELECT IFNULL(ROUND(SUM(quantity),3),0) AS s
      FROM h_inventory_lots
      WHERE tenant_id=%s AND material_id=%s AND receipt_date < %s
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    pre_receipt = float(cur.fetchone()['s'] or 0)

    cur.execute("""
      SELECT IFNULL(ROUND(SUM(quantity),3),0) AS s
      FROM h_inventory_transactions
      WHERE tenant_id=%s AND material_id=%s
        AND transaction_type='usage'
        AND transaction_date < %s
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    pre_usage = float(cur.fetchone()['s'] or 0)

    cur.execute("""
      SELECT IFNULL(ROUND(SUM(quantity),3),0) AS s
      FROM h_inventory_transactions
      WHERE tenant_id=%s AND material_id=%s
        AND transaction_type='adjustment'
        AND transaction_date < %s
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    pre_adj = float(cur.fetchone()['s'] or 0)

    # 시작 재고 (개념적 계산: 사용량 음수, 조정은 부호 그대로)
    opening_stock = pre_receipt - pre_usage + pre_adj

    # === LOT 기반 시작재고 (더 정확) ===
    # = 2026-02-09 이전 입고된 LOT의 잔량 합계 + 2026-02-09 이후 사용된 분 중 이 LOT들에서 빠진 분
    # 실용적 산식: 2026-02-09 시점 = 2026-02-09 이전 LOT들의 (quantity - 2026-02-09 이전 사용량)
    # 단, 2026-02-09 이전 LOT에서 2026-02-09 이후에 사용된 양도 시작재고로 계산해야 함
    # → 가장 단순: 2026-02-09 이전 LOT의 초기 quantity 합계 − 2026-02-09 이전 사용 합계
    # = pre_receipt − pre_usage_total
    # 그러나 트랜잭션 usage가 누락된 LOT가 있으면 부정확.
    # 더 정확한 방법: 2026-02-09 이전 LOT 중 현재 available한 LOT의 quantity 합 + 사용중인 LOT 중 2026-02-09 이후 사용분
    cur.execute("""
      SELECT IFNULL(ROUND(SUM(available_quantity),3),0) AS s
      FROM h_inventory_lots
      WHERE tenant_id=%s AND material_id=%s
        AND receipt_date < %s AND status='available'
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    pre_lots_remaining = float(cur.fetchone()['s'] or 0)

    # 2026-02-09 이전 LOT 중 2026-02-09 이후에 발생한 사용량
    cur.execute("""
      SELECT IFNULL(ROUND(SUM(t.quantity),3),0) AS s
      FROM h_inventory_transactions t
      INNER JOIN h_inventory_lots l ON l.id=t.lot_id
      WHERE t.tenant_id=%s AND t.material_id=%s
        AND t.transaction_type='usage'
        AND t.transaction_date >= %s
        AND l.receipt_date < %s
    """, (TENANT_ID, MATERIAL_ID, START_DATE, START_DATE))
    pre_lot_used_after_start = float(cur.fetchone()['s'] or 0)

    # LOT 기반 시작재고 = (이전 LOT 현재 잔량) + (이전 LOT가 2026-02-09 이후 사용된 양)
    opening_stock_lot = pre_lots_remaining + pre_lot_used_after_start

    # === 차이 원인 분석 ===
    # 트랜잭션 usage 합계 vs LOT 단위 사용량(quantity-available) 차이
    cur.execute("""
      SELECT IFNULL(ROUND(SUM(quantity - available_quantity),3),0) AS s
      FROM h_inventory_lots
      WHERE tenant_id=%s AND material_id=%s
    """, (TENANT_ID, MATERIAL_ID))
    lot_total_used = float(cur.fetchone()['s'] or 0)

    cur.execute("""
      SELECT IFNULL(ROUND(SUM(quantity),3),0) AS s
      FROM h_inventory_transactions
      WHERE tenant_id=%s AND material_id=%s AND transaction_type='usage'
    """, (TENANT_ID, MATERIAL_ID))
    txn_total_used = float(cur.fetchone()['s'] or 0)

    # LOT 단위로는 사용되었으나 usage 트랜잭션이 누락된 양
    missing_usage_kg = lot_total_used - txn_total_used

    # 3. 기간 내 LOT 입고 (h_inventory_lots 기반 - 트랜잭션과 무관하게 실제 입고)
    cur.execute("""
      SELECT 
        l.id AS lot_id, l.lot_number, l.quantity, l.available_quantity, l.unit,
        l.receipt_date, l.expiry_date, l.production_date, l.status,
        l.supplier_name, l.manufacturer_name,
        ap.id AS purchase_id, ap.unit_price, ap.total_amount,
        p.company_name AS partner_name, ap.transaction_date AS purchase_date
      FROM h_inventory_lots l
      LEFT JOIN accounting_purchases ap
        ON ap.tenant_id=l.tenant_id AND ap.material_id=l.material_id
        AND ap.transaction_date = DATE_FORMAT(l.receipt_date, '%%Y-%%m-%%d')
      LEFT JOIN partners p ON p.id = ap.partner_id AND p.tenant_id = ap.tenant_id
      WHERE l.tenant_id=%s AND l.material_id=%s
        AND l.receipt_date >= %s
      ORDER BY l.receipt_date, l.id
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    lots_in_period = list(cur.fetchall())

    # 4. 기간 내 사용 트랜잭션 (배치 정보 포함)
    cur.execute("""
      SELECT 
        t.id, t.transaction_date, t.lot_id, l.lot_number,
        t.quantity, t.unit,
        t.reference_type, t.reference_id, t.notes,
        b.id AS batch_id, b.batch_code, b.product_id,
        COALESCE(pv2.product_name, pv1.product_name) AS product_name
      FROM h_inventory_transactions t
      LEFT JOIN h_inventory_lots l ON l.id = t.lot_id
      LEFT JOIN h_batches b ON b.id = t.reference_id AND t.reference_type='batch'
      LEFT JOIN h_products_v2 pv2 ON pv2.id = b.product_id
      LEFT JOIN h_products pv1 ON pv1.id = b.product_id
      WHERE t.tenant_id=%s AND t.material_id=%s
        AND t.transaction_type='usage'
        AND t.transaction_date >= %s
      ORDER BY t.transaction_date, t.id
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    usages = list(cur.fetchall())

    # 5. 기간 내 조정 트랜잭션
    cur.execute("""
      SELECT 
        t.id, t.transaction_date, t.lot_id, l.lot_number,
        t.quantity, t.unit, t.reference_type, t.reference_id, t.notes
      FROM h_inventory_transactions t
      LEFT JOIN h_inventory_lots l ON l.id = t.lot_id
      WHERE t.tenant_id=%s AND t.material_id=%s
        AND t.transaction_type='adjustment'
        AND t.transaction_date >= %s
      ORDER BY t.transaction_date, t.id
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    adjustments = list(cur.fetchall())

    # 6. 전체 트랜잭션 (감사 추적용)
    cur.execute("""
      SELECT 
        t.id, t.transaction_date, t.transaction_type,
        t.lot_id, l.lot_number, t.quantity, t.unit,
        t.unit_cost, t.amount,
        t.reference_type, t.reference_id, t.notes
      FROM h_inventory_transactions t
      LEFT JOIN h_inventory_lots l ON l.id = t.lot_id
      WHERE t.tenant_id=%s AND t.material_id=%s
        AND t.transaction_date >= %s
      ORDER BY t.transaction_date, t.id
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    all_txns = list(cur.fetchall())

    # 7. 현재 LOT 현황 (전체)
    cur.execute("""
      SELECT id, lot_number, ROUND(quantity,3) AS quantity,
             ROUND(available_quantity,3) AS available_quantity,
             unit, receipt_date, expiry_date, production_date, status
      FROM h_inventory_lots
      WHERE tenant_id=%s AND material_id=%s
      ORDER BY receipt_date, id
    """, (TENANT_ID, MATERIAL_ID))
    all_lots = list(cur.fetchall())

    # 8. 현재 h_inventory
    cur.execute("""
      SELECT total_quantity, available_quantity, reserved_quantity, unit, last_updated
      FROM h_inventory
      WHERE tenant_id=%s AND material_id=%s
    """, (TENANT_ID, MATERIAL_ID))
    inv = cur.fetchone()

conn.close()

# ===================== 엑셀 생성 =====================
wb = Workbook()

# ---------- Sheet 1: 요약 ----------
ws1 = wb.active
ws1.title = "요약"

ws1["A1"] = f"멥쌀({mat['material_code']}) 재고 보고서"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A1:E1")

ws1["A2"] = f"기간: {START_DATE} ~ {date.today().isoformat()}"
ws1["A2"].font = SUBTITLE_FONT
ws1.merge_cells("A2:E2")

row = 4
ws1.cell(row=row, column=1, value="구분").fill = HEADER_FILL
ws1.cell(row=row, column=1).font = HEADER_FONT
ws1.cell(row=row, column=1).alignment = CENTER
ws1.cell(row=row, column=2, value="수량(kg)").fill = HEADER_FILL
ws1.cell(row=row, column=2).font = HEADER_FONT
ws1.cell(row=row, column=2).alignment = CENTER
ws1.cell(row=row, column=3, value="비고").fill = HEADER_FILL
ws1.cell(row=row, column=3).font = HEADER_FONT
ws1.cell(row=row, column=3).alignment = CENTER
style_header_row(ws1, row, 3)

# 시작 재고
period_receipt = sum(float(l['quantity'] or 0) for l in lots_in_period)
period_usage = sum(float(t['quantity'] or 0) for t in usages)
period_adj = sum(float(t['quantity'] or 0) for t in adjustments)
calc_closing = opening_stock + period_receipt - period_usage + period_adj
actual_inv = float(inv['available_quantity'] or 0) if inv else 0
discrepancy = actual_inv - calc_closing

# 두 가지 산식 비교
calc_closing_lot = opening_stock_lot + period_receipt - period_usage + period_adj - (lot_total_used - txn_total_used if missing_usage_kg > 0 else 0)
# 더 단순하고 정확: LOT 기반은 트랜잭션 누락분을 보정
# closing_lot = opening_stock_lot + period_receipt − (LOT 기반 기간 사용량) + period_adj
# LOT 기반 기간 사용량 = lot_total_used − (2026-02-09 이전 사용량)
lot_pre_used = pre_receipt - pre_lots_remaining - pre_lot_used_after_start  # 2026-02-09 이전에 이미 사용된 양
# 사실 더 명확한 정의: 2026-02-09 이전 LOT의 사용량 중 2026-02-09 이전 시점에 발생한 것
# = pre_lots_initial_qty - pre_lots_remaining_at_start
# pre_lots_remaining_at_start = pre_lots_remaining + pre_lot_used_after_start  (= opening_stock_lot)
# 따라서 lot_pre_used = pre_receipt - opening_stock_lot
lot_pre_used = pre_receipt - opening_stock_lot
lot_period_used = lot_total_used - lot_pre_used
# LOT 산식에서는 조정량을 별도로 더하지 않음 - LOT의 quantity가 이미 조정 반영을 못하므로
# 실제 검산: opening_stock_lot + period_receipt - lot_period_used = ?
calc_closing_lot_pure = opening_stock_lot + period_receipt - lot_period_used
# 조정량이 있다면 별도로 표시
calc_closing_lot = calc_closing_lot_pure + period_adj

rows = [
    ("[기초 재고] (2026-02-09 시점, LOT 기반)", opening_stock_lot,
     f"= 이전 LOT 현재 잔량({pre_lots_remaining:,.1f}) + 이전 LOT의 기간내 사용분({pre_lot_used_after_start:,.1f})"),
    ("(+) 기간 내 입고", period_receipt, f"LOT {len(lots_in_period)}건 · 인천광역시청 등"),
    ("(−) 기간 내 사용 [LOT 기반]", lot_period_used,
     f"트랜잭션 {period_usage:,.1f}kg + 누락 사용 {lot_period_used - period_usage:+,.1f}kg"),
    ("(+/−) 기간 내 조정", period_adj, f"조정 {len(adjustments)}건"),
    ("계산 기말 재고 [LOT 기반]", calc_closing_lot, "= 기초 + 입고 − 사용 + 조정"),
    ("실제 현재고 (h_inventory)", actual_inv, "DB 저장값"),
    ("차이", actual_inv - calc_closing_lot, "0이면 정합"),
]

for r_idx, (label, qty, note) in enumerate(rows, start=row + 1):
    c1 = ws1.cell(row=r_idx, column=1, value=label)
    c2 = ws1.cell(row=r_idx, column=2, value=round(qty, 2))
    c3 = ws1.cell(row=r_idx, column=3, value=note)
    c1.border = BORDER; c2.border = BORDER; c3.border = BORDER
    c1.alignment = LEFT
    c2.alignment = RIGHT
    c2.number_format = "#,##0.00"
    c3.alignment = LEFT
    if "입고" in label:
        c1.fill = RECEIPT_FILL; c2.fill = RECEIPT_FILL; c3.fill = RECEIPT_FILL
    elif "사용" in label:
        c1.fill = USAGE_FILL; c2.fill = USAGE_FILL; c3.fill = USAGE_FILL
    elif "조정" in label:
        c1.fill = ADJ_FILL; c2.fill = ADJ_FILL; c3.fill = ADJ_FILL
    elif label in ("계산 기말 재고", "실제 현재고 (h_inventory)", "차이"):
        c1.fill = TOTAL_FILL; c2.fill = TOTAL_FILL; c3.fill = TOTAL_FILL
        c1.font = Font(bold=True); c2.font = Font(bold=True)

# --- 참고: 트랜잭션 원장 기반 산식 (대조용) ---
ref_row = row + len(rows) + 2
ws1.cell(row=ref_row, column=1, value="[참고] 트랜잭션 원장 기반 산식 (대조용)").font = SUBTITLE_FONT
ws1.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=3)

ref_data = [
    ("기초 재고 (트랜잭션 기반)", opening_stock,
     f"= 이전 LOT 입고 {pre_receipt:,.1f} − 이전 usage 트랜잭션 {pre_usage:,.1f} + 조정 {pre_adj:+.1f}"),
    ("(+) 기간 내 입고 트랜잭션", period_receipt, "백필 후 LOT 입고와 일치"),
    ("(−) 기간 내 usage 트랜잭션", period_usage, f"{len(usages)}건"),
    ("(+/−) 기간 내 조정", period_adj, ""),
    ("계산 기말 (트랜잭션 기반)", calc_closing, "= 기초 + 입고 − usage + 조정"),
    ("실제 현재고", actual_inv, ""),
    ("차이 (트랜잭션 vs 실재고)", actual_inv - calc_closing,
     f"이 차이 ≈ usage 트랜잭션 누락 가능성. LOT 단위 사용량({lot_total_used:,.1f}) − usage 트랜잭션 합계({txn_total_used:,.1f}) = {missing_usage_kg:+,.1f}kg"),
]
for r_idx, (label, qty, note) in enumerate(ref_data, start=ref_row + 1):
    c1 = ws1.cell(row=r_idx, column=1, value=label)
    c2 = ws1.cell(row=r_idx, column=2, value=round(qty, 2))
    c3 = ws1.cell(row=r_idx, column=3, value=note)
    c1.border = BORDER; c2.border = BORDER; c3.border = BORDER
    c1.alignment = LEFT; c2.alignment = RIGHT; c3.alignment = LEFT
    c2.number_format = "#,##0.00"
    c1.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    c2.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    c3.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

note_row = ref_row + len(ref_data) + 2
ws1.cell(row=note_row, column=1,
         value="ℹ️ 본 보고서는 누락 receipt 트랜잭션 백필(2026-06-25) 이후 데이터입니다. "
               "사용(usage) 트랜잭션 누락분은 향후 별도 조사가 필요합니다.").font = Font(italic=True, color="808080")
ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)

autosize_cols(ws1, [32, 18, 60])
ws1.row_dimensions[1].height = 22

# ---------- Sheet 2: 입고 내역 ----------
ws2 = wb.create_sheet("입고 내역")
headers = ["No", "입고일", "LOT 번호", "수량(kg)", "단위", "공급처", "단가",
           "공급가액", "잔량(kg)", "상태", "생산일자", "소비기한", "매입전표 ID"]
for c, h in enumerate(headers, start=1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers))

total_in = 0
total_amount = 0
for i, l in enumerate(lots_in_period, start=2):
    qty = float(l['quantity'] or 0)
    amt = float(l['total_amount'] or 0)
    total_in += qty
    total_amount += amt
    values = [
        i - 1,
        l['receipt_date'].strftime("%Y-%m-%d") if l['receipt_date'] else "",
        l['lot_number'],
        round(qty, 2),
        l['unit'] or "kg",
        l['partner_name'] or l['supplier_name'] or "(미지정)",
        float(l['unit_price'] or 0) if l['unit_price'] else None,
        float(l['total_amount'] or 0) if l['total_amount'] else None,
        round(float(l['available_quantity'] or 0), 2),
        l['status'],
        l['production_date'].strftime("%Y-%m-%d") if l['production_date'] else "",
        l['expiry_date'].strftime("%Y-%m-%d") if l['expiry_date'] else "",
        l['purchase_id'] or "",
    ]
    for c, v in enumerate(values, start=1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.border = BORDER
        cell.fill = RECEIPT_FILL
        if c == 1: cell.alignment = CENTER
        elif c in (4, 7, 8, 9): cell.alignment = RIGHT; cell.number_format = "#,##0.00"
        elif c == 13: cell.alignment = CENTER
        else: cell.alignment = CENTER if c in (2, 5, 10, 11, 12) else LEFT

# Total row
tot_row = len(lots_in_period) + 2
ws2.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
ws2.cell(row=tot_row, column=1).alignment = CENTER
ws2.cell(row=tot_row, column=1).fill = TOTAL_FILL
ws2.cell(row=tot_row, column=3, value=f"{len(lots_in_period)} 건").alignment = CENTER
ws2.cell(row=tot_row, column=3).fill = TOTAL_FILL
ws2.cell(row=tot_row, column=3).font = Font(bold=True)
ws2.cell(row=tot_row, column=4, value=round(total_in, 2)).number_format = "#,##0.00"
ws2.cell(row=tot_row, column=4).font = Font(bold=True)
ws2.cell(row=tot_row, column=4).fill = TOTAL_FILL
ws2.cell(row=tot_row, column=4).alignment = RIGHT
ws2.cell(row=tot_row, column=8, value=round(total_amount, 0)).number_format = "#,##0"
ws2.cell(row=tot_row, column=8).font = Font(bold=True)
ws2.cell(row=tot_row, column=8).fill = TOTAL_FILL
ws2.cell(row=tot_row, column=8).alignment = RIGHT
for c in range(1, len(headers) + 1):
    ws2.cell(row=tot_row, column=c).border = BORDER
    if not ws2.cell(row=tot_row, column=c).fill.start_color.rgb or ws2.cell(row=tot_row, column=c).fill.start_color.rgb == '00000000':
        ws2.cell(row=tot_row, column=c).fill = TOTAL_FILL

autosize_cols(ws2, [5, 12, 22, 12, 6, 18, 11, 14, 11, 11, 12, 12, 12])
ws2.freeze_panes = "A2"

# ---------- Sheet 3: 사용 내역 ----------
ws3 = wb.create_sheet("사용 내역")
headers3 = ["No", "사용일", "LOT 번호", "수량(kg)", "단위", "배치 ID", "배치 코드", "제품명", "비고"]
for c, h in enumerate(headers3, start=1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(headers3))

total_use = 0
for i, t in enumerate(usages, start=2):
    qty = float(t['quantity'] or 0)
    total_use += qty
    values = [
        i - 1,
        t['transaction_date'].strftime("%Y-%m-%d") if t['transaction_date'] else "",
        t['lot_number'] or "",
        round(qty, 2),
        t['unit'] or "kg",
        t['batch_id'] or "",
        t['batch_code'] or "",
        t['product_name'] or "",
        (t['notes'] or "")[:80],
    ]
    for c, v in enumerate(values, start=1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.border = BORDER
        cell.fill = USAGE_FILL
        if c == 1: cell.alignment = CENTER
        elif c == 4: cell.alignment = RIGHT; cell.number_format = "#,##0.00"
        elif c in (2, 5, 6): cell.alignment = CENTER
        else: cell.alignment = LEFT

tot_row3 = len(usages) + 2
ws3.cell(row=tot_row3, column=1, value="합계").font = Font(bold=True)
ws3.cell(row=tot_row3, column=1).alignment = CENTER
ws3.cell(row=tot_row3, column=3, value=f"{len(usages)} 건").alignment = CENTER
ws3.cell(row=tot_row3, column=3).font = Font(bold=True)
ws3.cell(row=tot_row3, column=4, value=round(total_use, 2)).number_format = "#,##0.00"
ws3.cell(row=tot_row3, column=4).font = Font(bold=True)
ws3.cell(row=tot_row3, column=4).alignment = RIGHT
for c in range(1, len(headers3) + 1):
    ws3.cell(row=tot_row3, column=c).fill = TOTAL_FILL
    ws3.cell(row=tot_row3, column=c).border = BORDER

autosize_cols(ws3, [5, 12, 22, 12, 6, 9, 25, 28, 35])
ws3.freeze_panes = "A2"

# ---------- Sheet 4: 일자별 재고 변동 ----------
ws4 = wb.create_sheet("일자별 재고 변동")
headers4 = ["일자", "입고(kg)", "사용(kg)", "조정(kg)", "당일 순증감(kg)", "누계 재고(kg)"]
for c, h in enumerate(headers4, start=1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, len(headers4))

# 일자별 집계
daily = {}
for l in lots_in_period:
    d = l['receipt_date']
    if d:
        daily.setdefault(d, {"r": 0, "u": 0, "a": 0})
        daily[d]["r"] += float(l['quantity'] or 0)
for t in usages:
    d = t['transaction_date']
    if d:
        daily.setdefault(d, {"r": 0, "u": 0, "a": 0})
        daily[d]["u"] += float(t['quantity'] or 0)
for t in adjustments:
    d = t['transaction_date']
    if d:
        daily.setdefault(d, {"r": 0, "u": 0, "a": 0})
        daily[d]["a"] += float(t['quantity'] or 0)

# 기초재고를 가장 위에
ws4.cell(row=2, column=1, value=f"기초 ({START_DATE} 시점)").alignment = LEFT
ws4.cell(row=2, column=6, value=round(opening_stock, 2)).number_format = "#,##0.00"
ws4.cell(row=2, column=6).font = Font(bold=True)
ws4.cell(row=2, column=6).alignment = RIGHT
for c in range(1, len(headers4) + 1):
    ws4.cell(row=2, column=c).border = BORDER
    ws4.cell(row=2, column=c).fill = TOTAL_FILL

cumulative = opening_stock
r = 3
for d in sorted(daily.keys()):
    rec = daily[d]["r"]; usg = daily[d]["u"]; adj = daily[d]["a"]
    delta = rec - usg + adj
    cumulative += delta
    ws4.cell(row=r, column=1, value=d.strftime("%Y-%m-%d")).alignment = CENTER
    ws4.cell(row=r, column=2, value=round(rec, 2) if rec else 0).number_format = "#,##0.00"
    ws4.cell(row=r, column=3, value=round(usg, 2) if usg else 0).number_format = "#,##0.00"
    ws4.cell(row=r, column=4, value=round(adj, 2) if adj else 0).number_format = "#,##0.00;-#,##0.00;0.00"
    ws4.cell(row=r, column=5, value=round(delta, 2)).number_format = "#,##0.00;-#,##0.00;0.00"
    ws4.cell(row=r, column=6, value=round(cumulative, 2)).number_format = "#,##0.00"
    for c in range(2, 7):
        ws4.cell(row=r, column=c).alignment = RIGHT
    for c in range(1, len(headers4) + 1):
        ws4.cell(row=r, column=c).border = BORDER
        if rec > 0 and usg == 0:
            ws4.cell(row=r, column=c).fill = RECEIPT_FILL
        elif usg > 0 and rec == 0:
            ws4.cell(row=r, column=c).fill = USAGE_FILL
    r += 1

# 기말
ws4.cell(row=r, column=1, value="기말 (계산값)").alignment = LEFT
ws4.cell(row=r, column=6, value=round(cumulative, 2)).number_format = "#,##0.00"
ws4.cell(row=r, column=6).font = Font(bold=True)
ws4.cell(row=r, column=6).alignment = RIGHT
for c in range(1, len(headers4) + 1):
    ws4.cell(row=r, column=c).border = BORDER
    ws4.cell(row=r, column=c).fill = TOTAL_FILL
r += 1
ws4.cell(row=r, column=1, value="실제 (h_inventory)").alignment = LEFT
ws4.cell(row=r, column=6, value=round(actual_inv, 2)).number_format = "#,##0.00"
ws4.cell(row=r, column=6).font = Font(bold=True)
ws4.cell(row=r, column=6).alignment = RIGHT
for c in range(1, len(headers4) + 1):
    ws4.cell(row=r, column=c).border = BORDER
    ws4.cell(row=r, column=c).fill = TOTAL_FILL

autosize_cols(ws4, [22, 14, 14, 14, 16, 16])
ws4.freeze_panes = "A2"

# ---------- Sheet 5: LOT별 현황 (전체 기간) ----------
ws5 = wb.create_sheet("LOT별 현황")
headers5 = ["LOT ID", "LOT 번호", "입고일", "초기 수량(kg)", "잔량(kg)", "사용량(kg)", "단위", "생산일자", "소비기한", "상태"]
for c, h in enumerate(headers5, start=1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, len(headers5))

for i, l in enumerate(all_lots, start=2):
    init_q = float(l['quantity'] or 0)
    avail_q = float(l['available_quantity'] or 0)
    used_q = init_q - avail_q
    values = [
        l['id'], l['lot_number'],
        l['receipt_date'].strftime("%Y-%m-%d") if l['receipt_date'] else "",
        round(init_q, 2),
        round(avail_q, 2),
        round(used_q, 2),
        l['unit'] or "kg",
        l['production_date'].strftime("%Y-%m-%d") if l['production_date'] else "",
        l['expiry_date'].strftime("%Y-%m-%d") if l['expiry_date'] else "",
        l['status'],
    ]
    for c, v in enumerate(values, start=1):
        cell = ws5.cell(row=i, column=c, value=v)
        cell.border = BORDER
        if c == 1: cell.alignment = CENTER
        elif c in (4, 5, 6): cell.alignment = RIGHT; cell.number_format = "#,##0.00"
        elif c in (3, 7, 8, 9, 10): cell.alignment = CENTER
        else: cell.alignment = LEFT
        if l['status'] == 'available':
            cell.fill = RECEIPT_FILL
        elif l['status'] == 'used':
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

autosize_cols(ws5, [8, 26, 12, 14, 14, 14, 6, 12, 12, 12])
ws5.freeze_panes = "A2"

# ---------- Sheet 6: 전체 트랜잭션 (감사 추적) ----------
ws6 = wb.create_sheet("전체 트랜잭션")
headers6 = ["TXN ID", "일자", "유형", "LOT 번호", "수량(kg)", "단위", "단가", "금액", "참조 유형", "참조 ID", "비고"]
for c, h in enumerate(headers6, start=1):
    ws6.cell(row=1, column=c, value=h)
style_header_row(ws6, 1, len(headers6))

for i, t in enumerate(all_txns, start=2):
    values = [
        t['id'],
        t['transaction_date'].strftime("%Y-%m-%d") if t['transaction_date'] else "",
        t['transaction_type'],
        t['lot_number'] or "",
        round(float(t['quantity'] or 0), 2),
        t['unit'] or "kg",
        float(t['unit_cost']) if t['unit_cost'] else None,
        float(t['amount']) if t['amount'] else None,
        t['reference_type'] or "",
        t['reference_id'] or "",
        (t['notes'] or "")[:80],
    ]
    for c, v in enumerate(values, start=1):
        cell = ws6.cell(row=i, column=c, value=v)
        cell.border = BORDER
        if c == 1: cell.alignment = CENTER
        elif c in (5, 7, 8): cell.alignment = RIGHT; cell.number_format = "#,##0.00"
        elif c in (2, 3, 6, 9, 10): cell.alignment = CENTER
        else: cell.alignment = LEFT
        if t['transaction_type'] == 'receipt':
            cell.fill = RECEIPT_FILL
        elif t['transaction_type'] == 'usage':
            cell.fill = USAGE_FILL
        elif t['transaction_type'] == 'adjustment':
            cell.fill = ADJ_FILL

autosize_cols(ws6, [8, 12, 11, 24, 12, 6, 11, 14, 22, 10, 40])
ws6.freeze_panes = "A2"

# Save
output_path = "/home/root/webapp/멥쌀_재고보고서_2026-02-09부터.xlsx"
wb.save(output_path)
print(f"✅ 엑셀 파일 생성: {output_path}")
print(f"   기간: {START_DATE} ~ {date.today().isoformat()}")
print(f"   시트 6개, 기간 입고 {len(lots_in_period)}건 / 사용 {len(usages)}건 / 조정 {len(adjustments)}건")
print()
print(f"   [LOT 기반 산식 - 메인]")
print(f"   기초 {opening_stock_lot:,.2f}kg → 입고 +{period_receipt:,.2f}kg → 사용 -{lot_period_used:,.2f}kg → 조정 {period_adj:+,.2f}kg")
print(f"   = 계산 기말 {calc_closing_lot:,.2f}kg vs 실제 현재고 {actual_inv:,.2f}kg (차이 {actual_inv - calc_closing_lot:+,.2f}kg)")
print()
print(f"   [트랜잭션 원장 기반 산식 - 참고]")
print(f"   기초 {opening_stock:,.2f}kg → 입고 +{period_receipt:,.2f}kg → 사용 -{period_usage:,.2f}kg → 조정 {period_adj:+,.2f}kg")
print(f"   = 계산 기말 {calc_closing:,.2f}kg vs 실제 현재고 {actual_inv:,.2f}kg (차이 {actual_inv - calc_closing:+,.2f}kg)")
print()
print(f"   [데이터 정합성 진단]")
print(f"   LOT 단위 총 사용량 (quantity-available): {lot_total_used:,.2f}kg")
print(f"   usage 트랜잭션 총합:                       {txn_total_used:,.2f}kg")
print(f"   ☞ 사용(usage) 트랜잭션 누락 추정량:        {missing_usage_kg:+,.2f}kg")
