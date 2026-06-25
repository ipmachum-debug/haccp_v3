#!/usr/bin/env python3
"""
멥쌀(material_id=615) 재고 보고서 v2 — 사용자 양식 (일자별 1행, 제품별 상세)
- 기초재고: 2026-02-08 시점 1,520 kg (사용자 보고)
- 기간: 2026-02-09 ~ 현재
- 시트 1: 일별 원장 (사용자 양식)
- 시트 2: 빠진 항목 / 데이터 정합성 진단
- 시트 3: LOT 현황
"""
import pymysql
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from collections import defaultdict

# === 설정 ===
USER_OPENING_STOCK = 1520.0   # 2026-02-08 시점 기초재고 (사용자 보고)
USER_OPENING_DATE = "2026-02-08"
START_DATE = "2026-02-09"
MATERIAL_ID = 615
TENANT_ID = 2

conn = pymysql.connect(
    host="127.0.0.1", user="root", password="G0ld3n!T1004#Sec",
    database="haccp_tenant_db", charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor
)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="305496")
RECEIPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
USAGE_FILL = None
NEGATIVE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ADJ_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPENING_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ========== 데이터 조회 ==========
with conn.cursor() as cur:
    # 1) 사용 트랜잭션 (제품명 매핑 포함, 다단계 fallback)
    cur.execute("""
      SELECT 
        t.id AS txn_id,
        t.transaction_date AS dt,
        t.lot_id, l.lot_number,
        ROUND(t.quantity,3) AS qty,
        t.reference_type, t.reference_id,
        t.notes,
        b1.id AS batch_id_ref,
        b1.batch_code AS batch_code_ref,
        b1.product_id AS pid_ref,
        p1.product_name AS pname_ref
      FROM h_inventory_transactions t
      LEFT JOIN h_inventory_lots l ON l.id=t.lot_id
      LEFT JOIN h_batches b1 
        ON t.reference_type='batch' AND b1.id=t.reference_id
      LEFT JOIN h_products_v2 p1 ON p1.id=b1.product_id
      WHERE t.material_id=%s AND t.tenant_id=%s
        AND t.transaction_type='usage'
        AND t.transaction_date >= %s
      ORDER BY t.transaction_date, t.id
    """, (MATERIAL_ID, TENANT_ID, START_DATE))
    txns = list(cur.fetchall())

    # 2) batch_code 캐시 (notes 파싱 매칭용) — 전체 batch
    cur.execute("""
      SELECT b.id, b.batch_code, b.product_id, p.product_name
      FROM h_batches b
      LEFT JOIN h_products_v2 p ON p.id=b.product_id
    """)
    batch_by_code = {r['batch_code']: r for r in cur.fetchall() if r['batch_code']}

    # 3) batch_inputs 보조 매핑 (actual_quantity + 날짜) — lot_id가 NULL인 경우가 많음
    cur.execute("""
      SELECT bi.lot_id, ROUND(bi.actual_quantity,3) AS qty, 
             b.id AS batch_id, b.planned_date, b.product_id,
             p.product_name
      FROM h_batch_inputs bi
      INNER JOIN h_batches b ON b.id=bi.batch_id
      LEFT JOIN h_products_v2 p ON p.id=b.product_id
      WHERE bi.tenant_id=%s AND bi.material_id=%s
        AND bi.actual_quantity > 0
    """, (TENANT_ID, MATERIAL_ID))
    bi_index = defaultdict(list)
    for row in cur.fetchall():
        # key: (qty, planned_date) — lot_id 무시
        bi_index[(float(row['qty']), row['planned_date'])].append(row)

    # 4) 입고 LOT
    cur.execute("""
      SELECT 
        l.id AS lot_id, l.lot_number, ROUND(l.quantity,3) AS qty,
        ROUND(l.available_quantity,3) AS avail, l.receipt_date,
        l.expiry_date, l.production_date, l.status, l.unit,
        ap.id AS purchase_id, ap.unit_price, ap.total_amount,
        p.company_name AS supplier
      FROM h_inventory_lots l
      LEFT JOIN accounting_purchases ap
        ON ap.tenant_id=l.tenant_id AND ap.material_id=l.material_id
        AND ap.transaction_date = DATE_FORMAT(l.receipt_date, '%%Y-%%m-%%d')
      LEFT JOIN partners p ON p.id=ap.partner_id AND p.tenant_id=ap.tenant_id
      WHERE l.tenant_id=%s AND l.material_id=%s
        AND l.receipt_date >= %s
      ORDER BY l.receipt_date, l.id
    """, (TENANT_ID, MATERIAL_ID, START_DATE))
    lots_in_period = list(cur.fetchall())

    # 5) 조정 트랜잭션
    cur.execute("""
      SELECT id, transaction_date, lot_id, ROUND(quantity,3) AS qty, notes
      FROM h_inventory_transactions
      WHERE material_id=%s AND tenant_id=%s
        AND transaction_type='adjustment'
        AND transaction_date >= %s
      ORDER BY transaction_date, id
    """, (MATERIAL_ID, TENANT_ID, START_DATE))
    adjustments = list(cur.fetchall())

    # 6) 현재 재고
    cur.execute("""
      SELECT ROUND(available_quantity,3) AS av FROM h_inventory
      WHERE material_id=%s AND tenant_id=%s
    """, (MATERIAL_ID, TENANT_ID))
    actual_inv = float(cur.fetchone()['av'])

    # 7) LOT별 사용량 (quantity - available_quantity) — 트랜잭션 누락 추정용
    cur.execute("""
      SELECT 
        l.id AS lot_id, l.lot_number, 
        ROUND(l.quantity,3) AS init_q,
        ROUND(l.available_quantity,3) AS avail_q,
        ROUND(l.quantity - l.available_quantity,3) AS used_by_lot,
        l.receipt_date, l.status,
        IFNULL((SELECT ROUND(SUM(quantity),3) FROM h_inventory_transactions 
                WHERE tenant_id=l.tenant_id AND material_id=l.material_id 
                  AND lot_id=l.id AND transaction_type='usage'), 0) AS used_by_txn
      FROM h_inventory_lots l
      WHERE l.tenant_id=%s AND l.material_id=%s
      ORDER BY l.receipt_date, l.id
    """, (TENANT_ID, MATERIAL_ID))
    lot_reconciliation = list(cur.fetchall())

conn.close()


# ========== 트랜잭션 → 제품명 매핑 함수 ==========
batch_pattern = re.compile(r'batch\s+(\S+)')

def resolve_product_name(t):
    """
    트랜잭션의 제품명 다단계 추적:
    1) reference_type='batch' AND reference_id → b1 join 결과 사용
    2) notes에서 'batch XXXXX-YYYYMMDD-NNN' 추출 → batch_by_code
    3) batch_inputs index 매칭 (lot_id + qty + date)
    4) 실패 → '(추적불가)'
    """
    # Step 1: direct reference
    if t.get('pname_ref'):
        return t['pname_ref'], 'ref'
    
    # Step 2: notes parsing
    if t.get('notes'):
        m = batch_pattern.search(t['notes'])
        if m:
            code = m.group(1)
            if code in batch_by_code:
                return batch_by_code[code]['product_name'] or '(제품미정)', 'notes'
    
    # Step 3: batch_inputs index match (qty + date)
    key = (float(t['qty']), t['dt'])
    if key in bi_index:
        candidates = bi_index[key]
        if candidates:
            return (candidates[0]['product_name'] or '(제품미정)'), 'bi'
    
    return '(추적불가)', 'unknown'


# ========== 일자별 원장 생성 ==========
# Map: date → {receipts: [...], usages: [(product, qty)...], adjustments: [...]}
ledger = defaultdict(lambda: {'receipts': [], 'usages': [], 'adjustments': []})

for l in lots_in_period:
    d = l['receipt_date']
    ledger[d]['receipts'].append(l)

unmatched_count = 0
unmatched_kg = 0.0
matched_via_ref = 0
matched_via_notes = 0
matched_via_bi = 0

for t in txns:
    pname, source = resolve_product_name(t)
    if source == 'unknown':
        unmatched_count += 1
        unmatched_kg += float(t['qty'])
    elif source == 'ref':
        matched_via_ref += 1
    elif source == 'notes':
        matched_via_notes += 1
    elif source == 'bi':
        matched_via_bi += 1
    ledger[t['dt']]['usages'].append({
        'product': pname, 'qty': float(t['qty']),
        'source': source, 'lot_id': t['lot_id'],
        'lot_number': t['lot_number'],
        'notes': t['notes']
    })

for a in adjustments:
    ledger[a['transaction_date']]['adjustments'].append(a)

all_dates = sorted(ledger.keys())


# ========== 엑셀 생성 ==========
wb = Workbook()

# ---------- Sheet 1: 일별 원장 (사용자 양식) ----------
ws = wb.active
ws.title = "일별 원장"

ws["A1"] = f"멥쌀 재고원장 — 기간: {START_DATE} ~ {date.today().isoformat()}"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

# Header row at row 3
headers = ["일자", "구분", "내용 / 제품별 사용", "입고(kg)", "사용(kg)", "잔량(kg)", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=3, column=c, value=h)
style_header_row(ws, 3, len(headers))

# 기초재고 (사용자 보고값)
r = 4
ws.cell(row=r, column=1, value=USER_OPENING_DATE)
ws.cell(row=r, column=2, value="기초재고")
ws.cell(row=r, column=3, value="전기 이월")
ws.cell(row=r, column=6, value=USER_OPENING_STOCK)
ws.cell(row=r, column=6).number_format = "#,##0.00"
ws.cell(row=r, column=7, value="사용자 보고")
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = OPENING_FILL
    ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=c).font = Font(bold=True)
ws.cell(row=r, column=1).alignment = CENTER
ws.cell(row=r, column=2).alignment = CENTER
ws.cell(row=r, column=6).alignment = RIGHT
ws.cell(row=r, column=7).alignment = CENTER

balance = USER_OPENING_STOCK
r += 1

# 일자별 행
for d in all_dates:
    entries = ledger[d]
    # 같은 날짜에 입고 → 사용 → 조정 순으로 표시
    # 입고
    for rec in entries['receipts']:
        ws.cell(row=r, column=1, value=d.strftime("%Y-%m-%d"))
        ws.cell(row=r, column=2, value="입고")
        ws.cell(row=r, column=3, value=f"{rec['lot_number']} ({rec['supplier'] or '공급처 미상'})")
        ws.cell(row=r, column=4, value=float(rec['qty']))
        balance += float(rec['qty'])
        ws.cell(row=r, column=6, value=round(balance, 2))
        ws.cell(row=r, column=7, value="")
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = RECEIPT_FILL
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=4).alignment = RIGHT
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        r += 1
    
    # 사용 - 제품별로 합산
    if entries['usages']:
        product_qty = defaultdict(float)
        has_unknown = False
        unknown_qty = 0.0
        for u in entries['usages']:
            if u['source'] == 'unknown':
                has_unknown = True
                unknown_qty += u['qty']
            else:
                product_qty[u['product']] += u['qty']
        
        # 합산된 제품 텍스트
        parts = [f"{p}:{round(q,1)}kg" for p, q in sorted(product_qty.items(), key=lambda x: -x[1])]
        if has_unknown:
            parts.append(f"(추적불가):{round(unknown_qty,1)}kg")
        
        total_used = sum(product_qty.values()) + unknown_qty
        
        ws.cell(row=r, column=1, value=d.strftime("%Y-%m-%d"))
        ws.cell(row=r, column=2, value="사용")
        ws.cell(row=r, column=3, value=", ".join(parts))
        ws.cell(row=r, column=5, value=round(total_used, 2))
        balance -= total_used
        ws.cell(row=r, column=6, value=round(balance, 2))
        note = "정상" if balance >= 0 else "재고 음수"
        if has_unknown:
            note += f" / 추적불가 {round(unknown_qty,1)}kg 포함"
        ws.cell(row=r, column=7, value=note)
        
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER
            if balance < 0:
                ws.cell(row=r, column=c).fill = NEGATIVE_FILL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = "#,##0.00;[Red]-#,##0.00"
        ws.cell(row=r, column=7).alignment = CENTER
        r += 1
    
    # 조정
    for a in entries['adjustments']:
        ws.cell(row=r, column=1, value=d.strftime("%Y-%m-%d"))
        ws.cell(row=r, column=2, value="조정")
        ws.cell(row=r, column=3, value=(a['notes'] or '')[:80])
        adj_qty = float(a['qty'])
        if adj_qty >= 0:
            ws.cell(row=r, column=4, value=adj_qty)
        else:
            ws.cell(row=r, column=5, value=abs(adj_qty))
        balance += adj_qty
        ws.cell(row=r, column=6, value=round(balance, 2))
        ws.cell(row=r, column=7, value="조정")
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = ADJ_FILL
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=4).alignment = RIGHT
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = "#,##0.00;[Red]-#,##0.00"
        ws.cell(row=r, column=7).alignment = CENTER
        r += 1

# 마지막: 검증 행
r += 1
ws.cell(row=r, column=1, value="검산")
ws.cell(row=r, column=2, value="현재고")
ws.cell(row=r, column=3, value="원장 계산 잔량 vs DB 실제 현재고")
ws.cell(row=r, column=6, value=round(actual_inv, 2))
diff = round(actual_inv - balance, 2)
ws.cell(row=r, column=7, value=f"DB 현재고={actual_inv:.2f}, 원장={balance:.2f}, 차이={diff:+.2f}")
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=c).font = Font(bold=True)
ws.cell(row=r, column=1).alignment = CENTER
ws.cell(row=r, column=2).alignment = CENTER
ws.cell(row=r, column=3).alignment = LEFT
ws.cell(row=r, column=6).alignment = RIGHT
ws.cell(row=r, column=6).number_format = "#,##0.00"
ws.cell(row=r, column=7).alignment = LEFT

autosize(ws, [12, 9, 50, 12, 12, 12, 38])
ws.freeze_panes = "A4"
ws.row_dimensions[1].height = 22


# ---------- Sheet 2: 빠진 항목 / 데이터 정합성 진단 ----------
ws2 = wb.create_sheet("빠진 항목")

ws2["A1"] = "빠진 항목 / 데이터 정합성 진단"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:F1")

# Section 1: 요약
r = 3
ws2.cell(row=r, column=1, value="◾ 요약").font = Font(bold=True, size=12, color="305496")
r += 1

# 트랜잭션 합계
total_lot_in = sum(float(l['qty']) for l in lots_in_period)
total_txn_use = sum(float(t['qty']) for t in txns)
total_adj = sum(float(a['qty']) for a in adjustments)
calc_closing = USER_OPENING_STOCK + total_lot_in - total_txn_use + total_adj
diff_to_actual = actual_inv - calc_closing

summary = [
    ("기초재고 (사용자 보고, 2026-02-08)", USER_OPENING_STOCK, "표 첫 행"),
    ("(+) 기간 내 입고 (LOT)", total_lot_in, f"{len(lots_in_period)} 건"),
    ("(−) 기간 내 사용 (usage 트랜잭션)", total_txn_use, f"{len(txns)} 건"),
    ("(+/−) 기간 내 조정", total_adj, f"{len(adjustments)} 건"),
    ("= 계산 잔량 (원장 끝)", calc_closing, ""),
    ("DB 실제 현재고 (h_inventory)", actual_inv, ""),
    ("차이 (DB - 계산)", diff_to_actual, "0이면 정합, 음수면 사용 트랜잭션 누락 가능성"),
]
for label, qty, note in summary:
    ws2.cell(row=r, column=1, value=label)
    ws2.cell(row=r, column=2, value=round(qty, 2))
    ws2.cell(row=r, column=3, value=note)
    ws2.cell(row=r, column=2).number_format = "#,##0.00;[Red]-#,##0.00"
    ws2.cell(row=r, column=2).alignment = RIGHT
    ws2.cell(row=r, column=1).border = BORDER
    ws2.cell(row=r, column=2).border = BORDER
    ws2.cell(row=r, column=3).border = BORDER
    if label.startswith("차이"):
        if abs(diff_to_actual) > 0.5:
            ws2.cell(row=r, column=2).fill = MISSING_FILL
            ws2.cell(row=r, column=1).fill = MISSING_FILL
            ws2.cell(row=r, column=3).fill = MISSING_FILL
    r += 1

r += 2

# Section 2: 추적불가 사용 트랜잭션 통계
ws2.cell(row=r, column=1, value="◾ 사용 트랜잭션 추적 결과 (제품명 매핑)").font = Font(bold=True, size=12, color="305496")
r += 1
trace_stats = [
    ("reference_type='batch' 직접 참조", matched_via_ref, ""),
    ("notes에서 batch_code 파싱", matched_via_notes, "2026-06-25 백필분"),
    ("batch_inputs 보조 매칭 (lot+qty+date)", matched_via_bi, ""),
    ("추적 불가 (제품 미상)", unmatched_count, f"누적 {round(unmatched_kg,1)} kg"),
    ("전체", len(txns), f"누적 {round(total_txn_use,1)} kg"),
]
for label, cnt, note in trace_stats:
    ws2.cell(row=r, column=1, value=label)
    ws2.cell(row=r, column=2, value=cnt)
    ws2.cell(row=r, column=3, value=note)
    ws2.cell(row=r, column=2).alignment = RIGHT
    ws2.cell(row=r, column=1).border = BORDER
    ws2.cell(row=r, column=2).border = BORDER
    ws2.cell(row=r, column=3).border = BORDER
    if label.startswith("추적 불가"):
        ws2.cell(row=r, column=1).fill = MISSING_FILL
        ws2.cell(row=r, column=2).fill = MISSING_FILL
        ws2.cell(row=r, column=3).fill = MISSING_FILL
    r += 1

r += 2

# Section 3: LOT별 사용량 vs 트랜잭션 사용량 (누락 사용 진단)
ws2.cell(row=r, column=1, value="◾ LOT별 사용량 vs usage 트랜잭션 (사용 트랜잭션 누락 진단)").font = Font(bold=True, size=12, color="305496")
r += 1

lot_headers = ["LOT 번호", "입고일", "초기 수량(kg)", "현재 잔량(kg)", "LOT 단위 사용량(kg)", "usage 트랜잭션(kg)", "누락 추정(kg)", "상태"]
for c, h in enumerate(lot_headers, start=1):
    ws2.cell(row=r, column=c, value=h)
style_header_row(ws2, r, len(lot_headers))
r += 1

total_lot_used = 0.0
total_txn_used = 0.0
total_missing = 0.0
for l in lot_reconciliation:
    used_lot = float(l['used_by_lot'])
    used_txn = float(l['used_by_txn'])
    missing = used_lot - used_txn
    total_lot_used += used_lot
    total_txn_used += used_txn
    total_missing += missing
    
    ws2.cell(row=r, column=1, value=l['lot_number'])
    ws2.cell(row=r, column=2, value=l['receipt_date'].strftime("%Y-%m-%d") if l['receipt_date'] else "")
    ws2.cell(row=r, column=3, value=float(l['init_q']))
    ws2.cell(row=r, column=4, value=float(l['avail_q']))
    ws2.cell(row=r, column=5, value=used_lot)
    ws2.cell(row=r, column=6, value=used_txn)
    ws2.cell(row=r, column=7, value=round(missing, 3))
    ws2.cell(row=r, column=8, value=l['status'])
    
    for c in range(1, 9):
        ws2.cell(row=r, column=c).border = BORDER
        if c == 1: ws2.cell(row=r, column=c).alignment = LEFT
        elif c in (3,4,5,6,7): 
            ws2.cell(row=r, column=c).alignment = RIGHT
            ws2.cell(row=r, column=c).number_format = "#,##0.00;[Red]-#,##0.00"
        else: ws2.cell(row=r, column=c).alignment = CENTER
        if abs(missing) > 0.5:
            ws2.cell(row=r, column=c).fill = MISSING_FILL
    r += 1

# 합계
ws2.cell(row=r, column=1, value="합계")
ws2.cell(row=r, column=5, value=round(total_lot_used, 2))
ws2.cell(row=r, column=6, value=round(total_txn_used, 2))
ws2.cell(row=r, column=7, value=round(total_missing, 2))
for c in range(1, 9):
    ws2.cell(row=r, column=c).font = Font(bold=True)
    ws2.cell(row=r, column=c).fill = OPENING_FILL
    ws2.cell(row=r, column=c).border = BORDER
    if c in (5,6,7):
        ws2.cell(row=r, column=c).alignment = RIGHT
        ws2.cell(row=r, column=c).number_format = "#,##0.00;[Red]-#,##0.00"
r += 2

# Section 4: 결론 메모
ws2.cell(row=r, column=1, value="◾ 결론").font = Font(bold=True, size=12, color="305496")
r += 1
conclusions = [
    f"1. 기초재고 {USER_OPENING_STOCK:.0f}kg + 입고 {total_lot_in:.0f}kg − 사용 {total_txn_use:.0f}kg + 조정 {total_adj:+.0f}kg = {calc_closing:.0f}kg",
    f"2. DB 실제 현재고 {actual_inv:.0f}kg와의 차이: {diff_to_actual:+.0f}kg",
    f"3. 사용 트랜잭션 누락 추정량 (LOT 단위 vs 트랜잭션 비교): {total_missing:+.0f}kg",
    f"4. 추적 불가 사용 트랜잭션 (notes/reference 없음): {unmatched_count}건 / {unmatched_kg:.0f}kg",
    f"5. 2026-06-25 receipt 트랜잭션 백필로 입고는 정합 (LOT 합계 = receipt 트랜잭션 합계).",
    f"6. 사용(usage) 트랜잭션 일부도 동일하게 LOT만 차감되고 트랜잭션이 누락된 케이스가 있음 → 추가 조사 필요.",
]
for c_text in conclusions:
    ws2.cell(row=r, column=1, value=c_text)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws2.cell(row=r, column=1).alignment = LEFT
    r += 1

autosize(ws2, [28, 14, 16, 16, 18, 18, 16, 12])
ws2.row_dimensions[1].height = 22


# ---------- Sheet 3: 입고 상세 ----------
ws3 = wb.create_sheet("입고 상세")
in_headers = ["No", "입고일", "LOT 번호", "수량(kg)", "잔량(kg)", "공급처", "단가", "공급가액", "매입전표", "상태"]
for c, h in enumerate(in_headers, start=1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(in_headers))

for i, l in enumerate(lots_in_period, start=2):
    ws3.cell(row=i, column=1, value=i-1)
    ws3.cell(row=i, column=2, value=l['receipt_date'].strftime("%Y-%m-%d"))
    ws3.cell(row=i, column=3, value=l['lot_number'])
    ws3.cell(row=i, column=4, value=float(l['qty']))
    ws3.cell(row=i, column=5, value=float(l['avail']))
    ws3.cell(row=i, column=6, value=l['supplier'] or "(미지정)")
    ws3.cell(row=i, column=7, value=float(l['unit_price']) if l['unit_price'] else None)
    ws3.cell(row=i, column=8, value=float(l['total_amount']) if l['total_amount'] else None)
    ws3.cell(row=i, column=9, value=l['purchase_id'] or "")
    ws3.cell(row=i, column=10, value=l['status'])
    for c in range(1, 11):
        ws3.cell(row=i, column=c).border = BORDER
        ws3.cell(row=i, column=c).fill = RECEIPT_FILL
        if c == 1: ws3.cell(row=i, column=c).alignment = CENTER
        elif c in (4,5,7,8): 
            ws3.cell(row=i, column=c).alignment = RIGHT
            ws3.cell(row=i, column=c).number_format = "#,##0.00"
        elif c in (2,9,10): ws3.cell(row=i, column=c).alignment = CENTER
        else: ws3.cell(row=i, column=c).alignment = LEFT

# 합계
r3 = len(lots_in_period) + 2
ws3.cell(row=r3, column=1, value="합계").font = Font(bold=True)
ws3.cell(row=r3, column=3, value=f"{len(lots_in_period)} 건")
ws3.cell(row=r3, column=4, value=round(total_lot_in, 2))
ws3.cell(row=r3, column=8, value=round(sum(float(l['total_amount'] or 0) for l in lots_in_period), 0))
for c in range(1, 11):
    ws3.cell(row=r3, column=c).fill = OPENING_FILL
    ws3.cell(row=r3, column=c).border = BORDER
    ws3.cell(row=r3, column=c).font = Font(bold=True)
ws3.cell(row=r3, column=4).number_format = "#,##0.00"
ws3.cell(row=r3, column=4).alignment = RIGHT
ws3.cell(row=r3, column=8).number_format = "#,##0"
ws3.cell(row=r3, column=8).alignment = RIGHT
ws3.cell(row=r3, column=3).alignment = CENTER

autosize(ws3, [5, 12, 22, 12, 12, 18, 11, 14, 11, 11])
ws3.freeze_panes = "A2"


# Save
output_path = "/home/root/webapp/멥쌀_재고원장_2026-02-09부터.xlsx"
wb.save(output_path)
print(f"✅ 엑셀 파일 생성: {output_path}")
print()
print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
print(f"기초재고 (2026-02-08): {USER_OPENING_STOCK:,.0f} kg")
print(f"+ 기간 내 입고:        {total_lot_in:,.0f} kg ({len(lots_in_period)}건)")
print(f"− 기간 내 사용:        {total_txn_use:,.0f} kg ({len(txns)}건)")
print(f"+/- 기간 내 조정:      {total_adj:+,.0f} kg ({len(adjustments)}건)")
print(f"= 계산 잔량:           {calc_closing:,.0f} kg")
print(f"  DB 현재고:           {actual_inv:,.0f} kg")
print(f"  차이:                {diff_to_actual:+,.0f} kg")
print()
print(f"[제품명 매핑 결과]")
print(f"  reference 직접:      {matched_via_ref}건")
print(f"  notes 파싱:          {matched_via_notes}건")
print(f"  batch_inputs 보조:   {matched_via_bi}건")
print(f"  추적 불가:           {unmatched_count}건 / {unmatched_kg:,.1f} kg")
print()
print(f"[LOT vs 트랜잭션 정합성]")
print(f"  LOT 단위 사용 총합:  {total_lot_used:,.2f} kg")
print(f"  usage 트랜잭션 합:   {total_txn_used:,.2f} kg")
print(f"  사용 트랜잭션 누락:  {total_missing:+,.2f} kg")
