#!/usr/bin/env python3
"""
멥쌀 재고원장 v5 생성 스크립트
=================================

목적:
  v4 베이스라인(60행, 종료잔량 2,869.98kg)을 기반으로 BOM 기반 분배 사용 행을 추가하여
  최종 잔량을 정확히 1,669.98kg에 맞춘다. (총 분배량 = 1,200.00kg)

분배 대상:
  - h_mf_reports / h_mf_report_versions / h_mf_ingredients 의 BOM 데이터 사용
  - material_id = 168 (멥쌀, mf_ingredients ID 공간)
  - ACTIVE 제품 중, v4에 이미 반영된 15개(설기/쑥개떡/모시개떡/멥쌀가루) 제외
  - 기간: 2026-02-09 ~ 2026-06-12 (Option A: 4월 이전 배치 포함)
  - h_batches.actual_quantity × (BOM% / 100) = 멥쌀 소요량

알고리즘:
  1) v4 60행을 그대로 읽어옴 (잔량 2,869.98)
  2) DB 쿼리로 대상 배치 목록 확보 (planned_date DESC, id DESC)
  3) 같은 날짜는 제품별로 합산하여 "사용" 행 생성
  4) 가장 최근일부터 누적, 1,200kg에 도달하는 행은 부분 스케일링
  5) 새 사용 행들을 v4 타임라인의 올바른 위치에 삽입 (날짜 오름차순 유지)
  6) 잔량 재계산
  7) 최종 잔량이 1,669.98이 되는지 검증

Usage:
  python3 export_mepsal_inventory_report_v5.py
"""
import pymysql
from urllib.parse import urlparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from datetime import date
import os
import sys

# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "mysql://root:G0ld3n%21T1004%23Sec@127.0.0.1:3306/haccp_tenant_db",
)
TENANT_ID = 2
MEPSAL_MAT_ID_IN_MF = 168  # h_mf_ingredients.material_id 공간에서의 멥쌀

START_DATE = "2026-02-09"
END_DATE = "2026-06-12"
TARGET_DEPLETION = 1200.00  # kg, v4 종료잔량 2869.98 - 1200 = 1669.98

# v4에 이미 포함된 제품들 (제외 대상)
V4_PRODUCTS = {
    "꿀설기", "단호박설기", "딸기설기", "모듬설기", "모시개떡",
    "습식 멥쌀가루", "쑥 꿀설기", "쑥개떡", "우유설기", "자색고구마설기",
    "초코설기", "치즈설기", "카스테라설기", "호박 꿀설기", "흑임자설기",
}

BASE_FILE = "/home/root/webapp/멥쌀_재고원장_설기개떡멥쌀가루_수정전_v4.xlsx"
OUT_FILE = "/home/root/webapp/멥쌀_재고원장_설기개떡멥쌀가루_수정전_v5.xlsx"
EXPECTED_BASELINE_END = 2869.98
TARGET_FINAL_BALANCE = 1669.98


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def connect_db():
    u = urlparse(DATABASE_URL)
    password = (u.password or "").replace("%21", "!").replace("%23", "#")
    return pymysql.connect(
        host=u.hostname, port=u.port or 3306,
        user=u.username, password=password,
        database=u.path.lstrip("/"),
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )


def fetch_bom_map(conn):
    """ACTIVE 제품 중 멥쌀 BOM 가진 제품들의 v2_name -> 멥쌀 % 매핑."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT p2.product_name AS v2_name,
                   CAST(mi.quantity AS DECIMAL(10,4)) AS mepsal_pct
            FROM h_mf_reports mr
            JOIN h_products_v2 p2 ON p2.id = mr.product_id AND p2.is_active = 1
            JOIN h_mf_report_versions mrv ON mrv.mf_report_id = mr.id
            JOIN h_mf_ingredients mi ON mi.mf_report_version_id = mrv.id
            WHERE mr.tenant_id = %s
              AND mr.status = 'ACTIVE'
              AND mi.material_id = %s
              AND mrv.approval_status = 'APPROVED'
              AND mrv.id = (
                  SELECT MAX(id) FROM h_mf_report_versions
                  WHERE mf_report_id = mr.id AND approval_status = 'APPROVED'
              )
            ORDER BY p2.product_name
        """, (TENANT_ID, MEPSAL_MAT_ID_IN_MF))
        return {r["v2_name"]: float(r["mepsal_pct"]) for r in cur.fetchall()}


def fetch_target_batches(conn, target_names):
    """h_batches에서 대상 제품의 배치 조회 (최신 날짜 우선)."""
    if not target_names:
        return [], {}
    with conn.cursor() as cur:
        ph = ",".join(["%s"] * len(target_names))
        cur.execute(
            f"SELECT id, product_name FROM h_products "
            f"WHERE tenant_id=%s AND product_name IN ({ph})",
            (TENANT_ID, *target_names),
        )
        p_map = {r["product_name"]: r["id"] for r in cur.fetchall()}
        if not p_map:
            return [], {}

        ids = list(p_map.values())
        ph2 = ",".join(["%s"] * len(ids))
        cur.execute(
            f"SELECT b.id, b.batch_code, b.planned_date, p.product_name,"
            f" b.actual_quantity "
            f"FROM h_batches b "
            f"JOIN h_products p ON p.id = b.product_id "
            f"WHERE b.tenant_id=%s "
            f"  AND b.product_id IN ({ph2}) "
            f"  AND b.planned_date BETWEEN %s AND %s "
            f"  AND b.actual_quantity > 0 "
            f"ORDER BY b.planned_date DESC, b.id DESC",
            (TENANT_ID, *ids, START_DATE, END_DATE),
        )
        return cur.fetchall(), p_map


# ---------------------------------------------------------------------------
# 분배 계산
# ---------------------------------------------------------------------------
def round2(x: float) -> float:
    """소수 둘째 자리 반올림 (Excel 표기 일관성)."""
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def compute_distribution(batches, bom_map):
    """
    역연대순으로 누적, 1,200kg에 도달하는 시점을 찾고
    마지막 배치는 정확히 합계가 1,200이 되도록 스케일링.

    Returns:
      used_batches: [(planned_date, product_name, raw_qty, scale, mepsal_kg), ...]
                    (가장 최근일부터 정렬, 마지막이 스케일된 부분 사용)
      total: 1,200.00
    """
    used = []
    cumulative = 0.0
    target = TARGET_DEPLETION

    for b in batches:
        pct = bom_map.get(b["product_name"])
        if pct is None:
            continue
        qty = float(b["actual_quantity"])
        full_mepsal = round2(qty * pct / 100.0)
        if full_mepsal <= 0:
            continue

        remaining = round2(target - cumulative)
        if remaining <= 0:
            break

        if full_mepsal <= remaining:
            used.append({
                "date": b["planned_date"],
                "product": b["product_name"],
                "qty": qty,
                "pct": pct,
                "mepsal": full_mepsal,
                "scale": 1.0,
                "partial": False,
            })
            cumulative = round2(cumulative + full_mepsal)
        else:
            # 부분 스케일링 (마지막 배치)
            scale = remaining / full_mepsal
            used.append({
                "date": b["planned_date"],
                "product": b["product_name"],
                "qty": qty,
                "pct": pct,
                "mepsal": remaining,
                "scale": scale,
                "partial": True,
            })
            cumulative = round2(cumulative + remaining)
            break

    return used, cumulative


# ---------------------------------------------------------------------------
# v4 읽기 + v5 작성
# ---------------------------------------------------------------------------
def read_v4_rows():
    """v4 엑셀에서 헤더+데이터 행을 (date,kind,content,in_kg,use_kg,balance,note)로 반환."""
    wb = load_workbook(BASE_FILE)
    ws = wb.active
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows


def build_v5(v4_rows, used_batches, total_distributed):
    """
    v4 사용 행에 + 새 분배 사용 행을 날짜순으로 병합.
    같은 날짜 묶음은 한 행으로 합치되, v4 기존 행에 합쳐서는 안 됨(별도 신규 행).
    """
    # v4 헤더는 1행, 2행=기초재고
    header = v4_rows[0]
    opening = v4_rows[1]  # ['기초재고', ..., 1520, ...]

    # v4 데이터 행(3행 이후)을 dict 리스트로
    data_rows = []
    for raw in v4_rows[2:]:
        if not any(raw):
            continue
        data_rows.append({
            "date": str(raw[0]) if raw[0] else None,
            "kind": raw[1],
            "content": raw[2],
            "in_kg": raw[3],
            "use_kg": raw[4],
            "balance": None,  # 재계산
            "note": raw[6],
            "source": "v4",
        })

    # 같은 날짜의 분배 배치들을 제품별로 합산
    by_date = defaultdict(lambda: defaultdict(float))
    for u in used_batches:
        d = str(u["date"])
        by_date[d][u["product"]] = round2(
            by_date[d][u["product"]] + u["mepsal"]
        )

    # 새 사용 행 생성
    new_rows = []
    for d, prods in by_date.items():
        parts = [f"{name}:{kg:g}kg" for name, kg in prods.items()]
        total = round2(sum(prods.values()))
        new_rows.append({
            "date": d,
            "kind": "사용",
            "content": ", ".join(parts),
            "in_kg": None,
            "use_kg": total,
            "balance": None,
            "note": "BOM 기반 분배(v5)",
            "source": "v5_new",
        })

    # 병합 후 날짜 오름차순 정렬 (None=기초재고는 맨 앞에 따로 둠)
    merged = data_rows + new_rows

    def sort_key(r):
        d = r["date"]
        # v4 기존행과 v5 신규행이 같은 날짜에 있을 때:
        #   v4 기존행을 먼저, v5 신규행을 나중에 (source 순)
        priority = 0 if r["source"] == "v4" else 1
        return (d, priority)

    merged.sort(key=sort_key)

    # 잔량 재계산
    balance = float(opening[5])  # 기초재고 = 1520
    for r in merged:
        if r["kind"] == "입고":
            balance = round2(balance + float(r["in_kg"] or 0))
        elif r["kind"] == "사용":
            balance = round2(balance - float(r["use_kg"] or 0))
        r["balance"] = balance

    final_balance = balance

    return header, opening, merged, final_balance


def write_v5_excel(header, opening, merged_rows, final_balance, used_batches,
                   total_distributed):
    wb = Workbook()
    ws = wb.active
    ws.title = "일별 원장"

    # 스타일
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    in_fill = PatternFill("solid", fgColor="E2EFDA")
    use_fill = PatternFill("solid", fgColor="FFF2CC")
    new_use_fill = PatternFill("solid", fgColor="FCE4D6")  # v5 신규(주황)
    open_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # 헤더
    for i, h in enumerate(header, start=1):
        c = ws.cell(1, i, h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    # 기초재고
    for i, v in enumerate(opening, start=1):
        c = ws.cell(2, i, v)
        c.fill = open_fill
        c.font = Font(bold=True)
        c.border = border
        c.alignment = center if i in (1, 2, 7) else (right if i in (4, 5, 6) else left)

    # 데이터
    for idx, r in enumerate(merged_rows, start=3):
        ws.cell(idx, 1, r["date"]).alignment = center
        ws.cell(idx, 2, r["kind"]).alignment = center
        ws.cell(idx, 3, r["content"]).alignment = left
        ws.cell(idx, 4, r["in_kg"]).alignment = right
        ws.cell(idx, 5, r["use_kg"]).alignment = right
        ws.cell(idx, 6, r["balance"]).alignment = right
        ws.cell(idx, 7, r["note"]).alignment = center

        if r["kind"] == "입고":
            fill = in_fill
        elif r["source"] == "v5_new":
            fill = new_use_fill
        else:
            fill = use_fill

        for c in range(1, 8):
            cell = ws.cell(idx, c)
            cell.fill = fill
            cell.border = border
            if c in (4, 5, 6) and cell.value is not None:
                cell.number_format = "#,##0.00"

    # 컬럼 폭
    widths = [12, 10, 60, 12, 12, 14, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 22

    # ============ 분배 상세 시트 ============
    ws2 = wb.create_sheet("분배 상세 (v5 신규)")
    hdr2 = ["일자", "제품명", "생산량(kg)", "BOM 멥쌀(%)", "멥쌀 사용(kg)",
            "스케일", "비고"]
    for i, h in enumerate(hdr2, start=1):
        c = ws2.cell(1, i, h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    for idx, u in enumerate(used_batches, start=2):
        ws2.cell(idx, 1, str(u["date"])).alignment = center
        ws2.cell(idx, 2, u["product"]).alignment = left
        ws2.cell(idx, 3, u["qty"]).alignment = right
        ws2.cell(idx, 4, u["pct"]).alignment = right
        ws2.cell(idx, 5, u["mepsal"]).alignment = right
        ws2.cell(idx, 6, f"{u['scale']:.6f}").alignment = right
        ws2.cell(idx, 7, "부분 사용 (1,200kg 정확히 도달)" if u["partial"]
                 else "전체 사용").alignment = left
        for c in range(1, 8):
            cell = ws2.cell(idx, c)
            cell.border = border
            if c in (3, 4, 5):
                cell.number_format = "#,##0.0000" if c == 4 else "#,##0.00"

    # 합계 행
    total_row = len(used_batches) + 2
    ws2.cell(total_row, 1, "합계").font = Font(bold=True)
    ws2.cell(total_row, 1).alignment = center
    ws2.cell(total_row, 5, total_distributed).font = Font(bold=True)
    ws2.cell(total_row, 5).alignment = right
    ws2.cell(total_row, 5).number_format = "#,##0.00"
    for c in range(1, 8):
        ws2.cell(total_row, c).fill = open_fill
        ws2.cell(total_row, c).border = border

    widths2 = [12, 32, 14, 14, 14, 12, 32]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # ============ 요약 시트 ============
    ws3 = wb.create_sheet("요약")
    rows_summary = [
        ["항목", "값"],
        ["시작 재고 (2026-02-08)", 1520.00],
        ["v4 종료 잔량 (2026-06-25)", EXPECTED_BASELINE_END],
        ["v5 추가 분배 (역연대순)", -total_distributed],
        ["v5 최종 잔량 (2026-06-25)", final_balance],
        ["목표 잔량", TARGET_FINAL_BALANCE],
        ["차이", round2(final_balance - TARGET_FINAL_BALANCE)],
        ["분배 기간", f"{used_batches[-1]['date']} ~ {used_batches[0]['date']}"
            if used_batches else "-"],
        ["분배 배치 수", len(used_batches)],
        ["분배 제품 수", len(set(u["product"] for u in used_batches))],
    ]
    for r, row in enumerate(rows_summary, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws3.cell(r, c, val)
            cell.border = border
            if r == 1:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = center
            else:
                cell.alignment = left if c == 1 else right
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22

    wb.save(OUT_FILE)
    print(f"\n[OK] v5 저장: {OUT_FILE}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 78)
    print("멥쌀 재고원장 v5 생성 (Option A: 기간 확장)")
    print("=" * 78)
    conn = connect_db()
    try:
        bom_map = fetch_bom_map(conn)
        print(f"멥쌀 BOM 보유 ACTIVE 제품: {len(bom_map)}개")

        # v4 제외 후 대상
        target_names = [n for n in bom_map if n not in V4_PRODUCTS]
        print(f"분배 대상 후보 (v4 제외): {len(target_names)}개")

        batches, p_map = fetch_target_batches(conn, target_names)
        print(f"  - h_products 매핑 성공: {len(p_map)}개 제품")
        print(f"  - 매핑 안 됨: {len(target_names) - len(p_map)}개 "
              f"(h_batches 사용 안 함)")
        print(f"  - {START_DATE} ~ {END_DATE} 배치: {len(batches)}건")

        total_available = sum(
            float(b["actual_quantity"]) * bom_map[b["product_name"]] / 100.0
            for b in batches
        )
        print(f"  - 가용 BOM-멥쌀 총량: {total_available:.2f}kg")

        used_batches, total_distributed = compute_distribution(batches, bom_map)
        print(f"\n분배 결과:")
        print(f"  사용 배치: {len(used_batches)}건")
        print(f"  분배 합계: {total_distributed:.2f}kg (목표: {TARGET_DEPLETION}kg)")

        # 분배 명세 (역연대순, 최신→과거)
        print(f"\n분배 명세 (역연대순):")
        running = 0.0
        for u in used_batches:
            running = round2(running + u["mepsal"])
            tag = " <PARTIAL>" if u["partial"] else ""
            print(f"  {u['date']} | {u['product']:30s} | "
                  f"qty={u['qty']:>7.2f} × {u['pct']:>5.2f}% × {u['scale']:.4f} "
                  f"= {u['mepsal']:>7.2f}kg | 누적={running:>8.2f}kg{tag}")

    finally:
        conn.close()

    # v4 -> v5
    v4_rows = read_v4_rows()
    header, opening, merged, final_balance = build_v5(
        v4_rows, used_batches, total_distributed
    )

    print(f"\n[검증]")
    print(f"  v4 종료 잔량 (예상): {EXPECTED_BASELINE_END}")
    print(f"  v5 최종 잔량       : {final_balance}")
    print(f"  목표 잔량          : {TARGET_FINAL_BALANCE}")
    print(f"  차이               : {final_balance - TARGET_FINAL_BALANCE:+.2f}")

    if abs(final_balance - TARGET_FINAL_BALANCE) > 0.01:
        print("\n[ERROR] 최종 잔량이 목표와 일치하지 않음! 분배 계산을 재검토하세요.")
        sys.exit(1)

    write_v5_excel(header, opening, merged, final_balance, used_batches,
                   total_distributed)
    print(f"\n[OK] 최종 잔량 {final_balance}kg = 목표 {TARGET_FINAL_BALANCE}kg ✓")


if __name__ == "__main__":
    main()
