#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
멥쌀 재고원장 v7 - 실제 배치 생산량(actual_quantity) 사용

v6의 BOM% 역산 가짜 추정값을 폐기하고, DB h_batches.actual_quantity 를
일자+제품 단위로 그대로 가져와서 일별 원장에 박는다.

키 매칭:
  일별 원장 일자(YYYY-MM-DD) + 제품명 → DB h_batches (planned_date 동일, product_name 동일)
  의 actual_quantity 합계
"""
import os
import re
import sys
from copy import copy
import pymysql
from openpyxl import load_workbook

SRC = "/root/haccp_v3/멥쌀_재고원장_v5.xlsx"
DST = "/root/haccp_v3/멥쌀_재고원장_v7.xlsx"

DB_CFG = dict(host='127.0.0.1', port=3306, user='root', password='G0ld3n!T1004#Sec',
              database='haccp_tenant_db', charset='utf8mb4',
              cursorclass=pymysql.cursors.DictCursor)


def load_db_actual_map():
    """DB에서 (일자, 제품명) → (actual_quantity 합, 멥쌀사용 합, 배치코드들) 매핑."""
    conn = pymysql.connect(**DB_CFG)
    m = {}
    sql = """
        SELECT CAST(b.planned_date AS CHAR) AS d,
               COALESCE(p2.product_name, p1.product_name, '?') AS pname,
               SUM(b.actual_quantity)  AS actual_total,
               SUM(b.planned_quantity) AS planned_total,
               SUM(t.qty)              AS mepsal_used,
               GROUP_CONCAT(DISTINCT b.batch_code SEPARATOR ',') AS batch_codes,
               COUNT(DISTINCT b.id)    AS batch_cnt
        FROM h_batches b
        JOIN (
            SELECT reference_id AS bid, SUM(quantity) AS qty
            FROM h_inventory_transactions
            WHERE tenant_id = 2 AND material_id = 615
              AND reference_type = 'batch'
              AND transaction_type IN ('usage','outbound')
            GROUP BY reference_id
        ) t ON t.bid = b.id
        LEFT JOIN h_products_v2 p2 ON p2.id = b.product_id AND p2.tenant_id = b.tenant_id
        LEFT JOIN h_products    p1 ON p1.id = b.product_id AND p1.tenant_id = b.tenant_id
        WHERE b.tenant_id = 2
        GROUP BY b.planned_date, b.product_id
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        for r in cur.fetchall():
            d_str = str(r['d'])[:10]  # 'YYYY-MM-DD'
            key = (d_str, r['pname'].strip())
            m[key] = {
                'actual': float(r['actual_total'] or 0),
                'planned': float(r['planned_total'] or 0),
                'mepsal': float(r['mepsal_used'] or 0),
                'batches': r['batch_codes'] or '',
                'cnt': int(r['batch_cnt'] or 0),
            }
    conn.close()
    return m


def parse_content(content):
    """'쑥개떡:123.0kg, 모시개떡:82.0kg' → [('쑥개떡', 123.0), ...]"""
    if not content:
        return []
    items = []
    for token in re.split(r"[,;]\s*", str(content)):
        token = token.strip()
        if not token:
            continue
        m = re.match(r"^(.+?)\s*:\s*([0-9.]+)\s*kg?\s*$", token, re.IGNORECASE)
        if m:
            try:
                items.append((m.group(1).strip(), float(m.group(2))))
            except ValueError:
                pass
    return items


def main():
    if not os.path.exists(SRC):
        sys.exit(f"ERROR: {SRC} 없음")

    print(f"[1] v5 로드: {SRC}")
    wb = load_workbook(SRC)

    # 분배 상세 시트의 (일자, 제품) → 생산량 보조 매핑 (DB 매칭 안되는 행용)
    dist_map = {}
    if "분배 상세 (v5 신규)" in wb.sheetnames:
        ws_d = wb["분배 상세 (v5 신규)"]
        for row in ws_d.iter_rows(min_row=2, values_only=True):
            d, name, prod_q, *_ = row + (None, None)
            if not d or not name or name == "합계":
                continue
            k = (str(d)[:10], str(name).strip())
            dist_map[k] = dist_map.get(k, 0) + (float(prod_q) if prod_q else 0)
        print(f"    분배상세 보조 매핑 {len(dist_map)}건 (DB 미존재 행용)")

    print("[2] DB 실제 생산량 로드 (h_batches.actual_quantity)")
    db_map = load_db_actual_map()
    print(f"    DB (일자,제품) 매핑 {len(db_map)}건")
    print("    샘플:")
    for k in sorted(db_map.keys())[:5]:
        v = db_map[k]
        print(f"      {k[0]} {k[1]:20s}  actual={v['actual']:>7.2f}  멥쌀={v['mepsal']:>7.2f}  배치{v['cnt']}건")
    for k in sorted(db_map.keys())[-3:]:
        v = db_map[k]
        print(f"      {k[0]} {k[1]:20s}  actual={v['actual']:>7.2f}  멥쌀={v['mepsal']:>7.2f}  배치{v['cnt']}건")

    print("[3] 일별 원장 시트 변환")
    ws = wb["일별 원장"]
    # G열에 새 컬럼 삽입 (기존 G=비고 → H로 밀림)
    ws.insert_cols(7)
    ws.cell(row=1, column=7, value="제품 생산량(kg)")

    # 헤더 스타일 복사
    hdr_a = ws.cell(row=1, column=1)
    new_hdr = ws.cell(row=1, column=7)
    if hdr_a.font:      new_hdr.font = copy(hdr_a.font)
    if hdr_a.fill:      new_hdr.fill = copy(hdr_a.fill)
    if hdr_a.alignment: new_hdr.alignment = copy(hdr_a.alignment)
    if hdr_a.border:    new_hdr.border = copy(hdr_a.border)

    matched = 0
    unmatched = []
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        kind = ws.cell(row=r, column=2).value
        content = ws.cell(row=r, column=3).value
        if kind != "사용" or not content:
            continue

        date_str = str(date_cell)[:10] if date_cell else ""
        items = parse_content(content)
        if not items:
            continue

        new_parts = []
        total_actual = 0.0
        all_batch_codes = []
        sources_used = []
        any_match = False
        for name, mepsal_kg in items:
            info = db_map.get((date_str, name))
            if info and info['actual'] > 0:
                actual = info['actual']
                total_actual += actual
                all_batch_codes.append(info['batches'])
                new_parts.append(f"{name}:{mepsal_kg}kg→생산 {actual:g}kg")
                sources_used.append("DB")
                any_match = True
            else:
                # DB에 없으면 분배 상세 시트(v5) 값 사용
                fb = dist_map.get((date_str, name))
                if fb and fb > 0:
                    total_actual += fb
                    new_parts.append(f"{name}:{mepsal_kg}kg→생산 {fb:g}kg")
                    sources_used.append("v5분배")
                    any_match = True
                else:
                    new_parts.append(f"{name}:{mepsal_kg}kg→생산 ?")
                    unmatched.append((date_str, name, mepsal_kg))

        ws.cell(row=r, column=3, value=", ".join(new_parts))
        if total_actual > 0:
            ws.cell(row=r, column=7, value=round(total_actual, 2))
            matched += 1

        # 비고에 출처/배치코드 추가
        note_cell = ws.cell(row=r, column=8)  # 새 G삽입으로 비고는 H로
        old_note = note_cell.value or ""
        if any_match:
            tag_parts = []
            if all_batch_codes:
                codes = ",".join(c for c in all_batch_codes if c)
                if codes:
                    tag_parts.append(f"배치:{codes}")
            uniq_src = ",".join(sorted(set(sources_used)))
            if uniq_src:
                tag_parts.append(f"출처:{uniq_src}")
            if tag_parts:
                tag = " [" + " | ".join(tag_parts) + "]"
                if tag.strip() not in (old_note or ""):
                    note_cell.value = f"{old_note}{tag}".strip()

    print(f"    매칭 성공: {matched} 행")
    if unmatched:
        print(f"    매칭 실패: {len(unmatched)} 건 (DB에 해당 일자+제품 배치 없음)")
        for u in unmatched[:10]:
            print(f"      - {u[0]} {u[1]} (멥쌀 {u[2]}kg)")
        if len(unmatched) > 10:
            print(f"      ... 외 {len(unmatched)-10}건")

    # 열 폭
    for col, w in {"A":12,"B":8,"C":75,"D":10,"E":10,"F":11,"G":15,"H":45}.items():
        ws.column_dimensions[col].width = w

    # 요약 시트 메타
    if "요약" in wb.sheetnames:
        ws_sum = wb["요약"]
        last = ws_sum.max_row + 1
        ws_sum.cell(row=last,   column=1, value="v7 보강")
        ws_sum.cell(row=last,   column=2, value="DB h_batches.actual_quantity 실제 생산량 사용")
        ws_sum.cell(row=last+1, column=1, value="생성일")
        ws_sum.cell(row=last+1, column=2, value="2026-06-30")
        ws_sum.cell(row=last+2, column=1, value="매칭 성공")
        ws_sum.cell(row=last+2, column=2, value=f"{matched} 행")
        ws_sum.cell(row=last+3, column=1, value="매칭 실패")
        ws_sum.cell(row=last+3, column=2, value=f"{len(unmatched)} 건")

    print(f"[4] v7 저장: {DST}")
    wb.save(DST)
    print("    완료")


if __name__ == "__main__":
    main()
